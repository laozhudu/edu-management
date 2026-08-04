"""
ExcelTemplateService 测试（Sprint 4.8.1）
覆盖：单值填充、行区块展开、样式保留、bytes 输出、模板保存
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from openpyxl import Workbook, load_workbook

from edu_system.services.report_excel import ExcelTemplateService


@pytest.fixture
def simple_template(tmp_path):
    """简单模板：A1=姓名{{name}} B1=分数{{score}}"""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "姓名：{{name}}"
    ws["B1"] = "分数：{{score}}"
    path = tmp_path / "tpl.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def rows_template(tmp_path):
    """行模板：第1行表头，第2行是 {{#rows}} 占位（实际用 {{name}}/{{score}}）"""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "姓名"
    ws["B1"] = "分数"
    ws["A2"] = "{{name}}"
    ws["B2"] = "{{score}}"
    path = tmp_path / "rows.xlsx"
    wb.save(path)
    return path


class TestFillCells:
    def test_fill_single_values(self, simple_template):
        data = {"name": "张三", "score": "95"}
        wb = load_workbook(str(simple_template))
        ws = wb.active
        filled = ExcelTemplateService.fill_cells(ws, data)
        assert filled == 2
        assert ws["A1"].value == "姓名：张三"
        assert ws["B1"].value == "分数：95"

    def test_partial_fill(self, simple_template):
        """只填部分键，未匹配的保留占位符"""
        data = {"name": "李四"}
        wb = load_workbook(str(simple_template))
        ws = wb.active
        filled = ExcelTemplateService.fill_cells(ws, data)
        assert filled == 1
        assert ws["A1"].value == "姓名：李四"
        assert ws["B1"].value == "分数：{{score}}"  # 未填充保留


class TestExpandRows:
    def test_expand_rows(self, rows_template):
        """行区块展开为多行数据"""
        wb = load_workbook(str(rows_template))
        ws = wb.active
        rows = [
            {"name": "张三", "score": 95},
            {"name": "李四", "score": 88},
        ]
        written = ExcelTemplateService.expand_rows(ws, rows, start_row=2, template_row=2)
        assert written == 2
        assert ws["A2"].value == "张三"
        assert ws["B2"].value == 95
        assert ws["A3"].value == "李四"
        assert ws["B3"].value == 88


class TestRender:
    def test_render_to_bytes(self, simple_template):
        data = {"name": "王五", "score": "90"}
        out = ExcelTemplateService.render(str(simple_template), data)
        assert isinstance(out, bytes)
        assert len(out) > 0

    def test_render_to_file(self, simple_template, tmp_path):
        data = {"name": "赵六", "score": "87"}
        out_path = tmp_path / "out.xlsx"
        ExcelTemplateService.render(str(simple_template), data, output_path=out_path)
        assert out_path.exists()
        # 验证渲染结果保留
        wb = load_workbook(str(out_path))
        assert wb.active["A1"].value == "姓名：赵六"
