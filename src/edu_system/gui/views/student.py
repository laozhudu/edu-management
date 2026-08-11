"""
GUI 视图 — 学生管理 (PyQt5 内存模式)
核心设计：
  1. _all_data  = 从DB加载的全部学生（只加载一次）
  2. _data      = 当前显示的数据（_all_data 的子集/排序结果）
  3. _dirty_ids = 被修改过的学生ID集合
  4. _deleted_ids = 标记删除的学生ID集合
  5. _new_students = 新增未保存的学生对象列表
  6. 筛选/排序/搜索 → 仅操作 _data，不碰数据库
  7. 编辑/新增/删除 → 标记 _dirty/_deleted/_new，保存按钮才批量写库
"""

from PyQt5.QtCore import QByteArray, Qt, QTimer
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSplitter,
    QTableView,
    QTextEdit,
    QToolButton,
    QVBoxLayout,
    QWidget,
)
from sqlalchemy.orm import Session

from edu_system.core import Permission, can
from edu_system.gui.theme import C, font
from edu_system.gui.views.base import ColumnSelectorDialog
from edu_system.gui.views.student_edit_dialog import StudentEditDialog  # noqa: F401
from edu_system.models import Class as ClassModel
from edu_system.models import Student


def _btn(txt, color, w=None):
    from edu_system.gui.components import btn

    return btn(txt, color, w)

ALL_COLUMNS = {
    "class_name": "班级",
    "student_no": "座号",
    "name": "姓名",
    "gender": "性别",
    "student_code": "学籍号",
    "id_card": "身份证",
    "phone": "电话",
    "status": "状态",
    "birth_date": "出生日期",
    "ethnicity": "民族",
    "native_place": "籍贯",
    "political_status": "政治面貌",
    "address": "居住地址",
    "hukou_addr": "户籍地址",
    "boarding": "走住读",
    "exam_no": "考号",
    "enroll_year": "入学年",
    "guardian1_name": "监护人1",
    "guardian1_relation": "监护人1关系",
    "guardian1_phone": "监护人1电话",
    "guardian1_work": "监护人1单位",
    "guardian1_edu": "监护人1学历",
    "guardian1_id_card": "监护人1身份证",
    "guardian2_name": "监护人2",
    "guardian2_relation": "监护人2关系",
    "guardian2_phone": "监护人2电话",
    "guardian2_work": "监护人2单位",
    "guardian2_edu": "监护人2学历",
    "guardian2_id_card": "监护人2身份证",
}
DEFAULT_COLUMNS = [
    "class_name",
    "student_no",
    "name",
    "gender",
    "student_code",
    "id_card",
    "hukou_addr",
    "phone",
    "status",
]
EDITABLE_FIELDS = [
    ("student_no", "座号"),
    ("name", "姓名"),
    ("gender", "性别"),
    ("student_code", "学籍号"),
    ("id_card", "身份证"),
    ("birth_date", "出生日期"),
    ("ethnicity", "民族"),
    ("native_place", "籍贯"),
    ("political_status", "政治面貌"),
    ("phone", "电话"),
    ("address", "居住地址"),
    ("hukou_addr", "户籍地址"),
    ("boarding", "走住读"),
    ("exam_no", "考号"),
    ("enroll_year", "入学年"),
    ("guardian1_name", "监护人1"),
    ("guardian1_relation", "监护人1关系"),
    ("guardian1_phone", "监护人1电话"),
    ("guardian1_work", "监护人1单位"),
    ("guardian1_edu", "监护人1学历"),
    ("guardian1_id_card", "监护人1身份证"),
    ("guardian2_name", "监护人2"),
    ("guardian2_relation", "监护人2关系"),
    ("guardian2_phone", "监护人2电话"),
    ("guardian2_work", "监护人2单位"),
    ("guardian2_edu", "监护人2学历"),
    ("guardian2_id_card", "监护人2身份证"),
    ("status", "状态"),
    ("note", "备注"),
]

# 导入字段映射（Excel列名标准化 → Student字段）
_IMPORT_FIELD_MAP = {
    "姓名": "name",
    "性别": "gender",
    "学籍号": "student_code",
    "身份证": "id_card",
    "身份证件号": "id_card",
    "电话": "phone",
    "联系电话": "phone",
    "民族": "ethnicity",
    "籍贯": "native_place",
    "政治面貌": "political_status",
    "居住地址": "address",
    "户籍地址": "hukou_addr",
    "现居住地址详细地址": "address",
    "户口所在地具体地址": "hukou_addr",
    "走住读": "boarding",
    "是否寄宿生": "boarding",
    "监护人1": "guardian1_name",
    "监护人1电话": "guardian1_phone",
    "监护人2": "guardian2_name",
    "监护人2电话": "guardian2_phone",
    "班级": "class_name",
}


class StudentView(QWidget):
    def __init__(self, session: Session):
        super().__init__()
        self.session = session
        self._all_data = []
        self._data = []
        self._visible_columns = self._load_column_prefs()
        self._sort_col, self._sort_asc = self._load_sort_prefs()
        self._dirty_ids = set()
        self._deleted_ids = set()
        self._new_students = []
        self._search_timer = None
        self._grade_cb = self._status_cb = self._class_cb = self._search = None
        self._page_size = 200
        self._current_page = 1
        self._total_count = 0
        self._loading = False

        # 初始化内存仓库
        from src.edu_system.services.memory_student import MemoryStudentRepository

        self._repo = MemoryStudentRepository(session)

        self._build_ui()
        self._setup_table_enhancements()
        QTimer.singleShot(0, self._load_first_page)

    def _school_name(self) -> str:
        """从 ui_config 读取校名（可灵活配置修改）"""
        try:
            from edu_system.config.ui_config import get_config

            cfg = get_config()
            return getattr(cfg, "school_name", "") or "示例学校"
        except Exception:
            return "示例学校"

    def refresh(self):
        self._load_first_page()

    # ═══════════════════════════════════
    # 偏好持久化
    # ═══════════════════════════════════

    def _load_column_prefs(self):
        from edu_system.models import Setting

        val = self.session.query(Setting).filter_by(key="student_view_columns").first()
        if val and val.value:
            cols = val.value.split(",")
            return [c for c in cols if c in ALL_COLUMNS]
        return list(DEFAULT_COLUMNS)

    def _save_column_prefs(self):
        from edu_system.models import Setting

        val = ",".join(self._visible_columns)
        entry = self.session.query(Setting).filter_by(key="student_view_columns").first()
        if entry:
            entry.value = val
        else:
            self.session.add(Setting(key="student_view_columns", value=val))
        self.session.commit()

    def _load_sort_prefs(self):
        from edu_system.models import Setting

        col_val = self.session.query(Setting).filter_by(key="student_view_sort_col").first()
        order_val = self.session.query(Setting).filter_by(key="student_view_sort_order").first()
        sort_col = int(col_val.value) if col_val and col_val.value.isdigit() else -1
        sort_asc = order_val.value == "asc" if order_val else True
        return sort_col, sort_asc

    def _save_sort_prefs(self, sort_col: int, sort_asc: bool):
        from edu_system.models import Setting

        for key, val in [
            ("student_view_sort_col", str(sort_col)),
            ("student_view_sort_order", "asc" if sort_asc else "desc"),
        ]:
            entry = self.session.query(Setting).filter_by(key=key).first()
            if entry:
                entry.value = val
            else:
                self.session.add(Setting(key=key, value=val))
        self.session.commit()

    # ═══════════════════════════════════
    # UI 构建
    # ═══════════════════════════════════

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 6, 8, 6)
        layout.setSpacing(4)

        # 概览条
        self._overview = QLabel()
        self._overview.setFont(font(9))
        self._overview.setStyleSheet(
            "background:white; border:1px solid #DDD; border-radius:3px; padding:5px 8px;"
        )
        layout.addWidget(self._overview)

        # 工具栏 - 分级：高频显性 + 批量下拉 + 更多次级
        tb = QHBoxLayout()
        tb.setSpacing(4)

        toolbar_items = [
            (
                "新增",
                C["accent_green"],
                self._add,
                "Ctrl+N",
                "新增学生 (Ctrl+N)",
                Permission.STUDENT_CREATE,
            ),
            (
                "导入",
                C["accent_blue"],
                self._import,
                "Ctrl+I",
                "导入学生名单 (Ctrl+I)",
                Permission.STUDENT_IMPORT,
            ),
            (
                "保存",
                C["accent_teal"],
                self._save,
                "Ctrl+S",
                "保存所有更改 (Ctrl+S)",
                Permission.STUDENT_EDIT,
            ),
            (
                "删除",
                C["accent_red"],
                self._delete_students,
                "Del",
                "删除选中学生 (Del)",
                Permission.STUDENT_DELETE,
            ),
            (
                "刷新",
                C["accent_orange"],
                self._load_first_page,
                "F5",
                "刷新数据 (F5)",
                Permission.STUDENT_VIEW,
            ),
            (
                "搜索",
                C["accent_purple"],
                lambda: self._search.setFocus(),
                "Ctrl+F",
                "聚焦搜索框 (Ctrl+F)",
                Permission.STUDENT_VIEW,
            ),
        ]

        for txt, clr, cb, shortcut, tooltip, perm in toolbar_items:
            if not can(perm):
                continue
            b = _btn(f"{txt} {shortcut}", clr)
            b.clicked.connect(cb)
            b.setToolTip(tooltip)
            tb.addWidget(b)

        tb.addSpacing(8)

        # 批量操作下拉
        batch_btn = QToolButton()
        batch_btn.setText("批量操作 ▼")
        batch_btn.setPopupMode(QToolButton.InstantPopup)
        batch_btn.setCursor(Qt.PointingHandCursor)
        batch_btn.setStyleSheet(
            f"QToolButton {{ background: {C['accent_orange']}; color: white; border: none; border-radius: 3px; padding: 4px 12px; font-size: 9pt; }} QToolButton:hover {{ background: #E67E22; }} QToolButton::menu-indicator {{ image: none; }}"
        )
        batch_menu = QMenu(self)
        for name, callback, perm in [
            ("转班", self._batch_transfer, Permission.STUDENT_BATCH),
            ("改状态", self._batch_change_status, Permission.STUDENT_BATCH),
            ("生成考号", self._generate_exam_nos, Permission.STUDENT_EDIT),
            ("重排座号", self._resort_seats, Permission.STUDENT_EDIT),
            ("导出选中", self._export_selected, Permission.STUDENT_EXPORT),
            ("打印名册", self._print_roster, Permission.STUDENT_EXPORT),
        ]:
            if can(perm):
                batch_menu.addAction(name, callback)
        batch_btn.setMenu(batch_menu)
        tb.addWidget(batch_btn)

        tb.addSpacing(8)

        # 更多操作下拉
        more_btn = QToolButton()
        more_btn.setText("更多 ▼")
        more_btn.setPopupMode(QToolButton.InstantPopup)
        more_btn.setCursor(Qt.PointingHandCursor)
        more_btn.setStyleSheet(
            "QToolButton { background: gray; color: white; border: none; border-radius: 3px; padding: 4px 12px; font-size: 9pt; } QToolButton:hover { background: #34495E; } QToolButton::menu-indicator { image: none; }"
        )
        more_menu = QMenu(self)
        for name, callback, perm in [
            ("导出", self._export, Permission.STUDENT_EXPORT),
            ("列选择", self._column_selector, Permission.STUDENT_VIEW),
            ("刷新", self._load_first_page, Permission.STUDENT_VIEW),
            ("补充信息", self._import_supplement, Permission.STUDENT_IMPORT),
        ]:
            if can(perm):
                more_menu.addAction(name, callback)
        more_btn.setMenu(more_menu)
        tb.addWidget(more_btn)

        tb.addStretch()

        self._status_lbl = QLabel()
        self._status_lbl.setFont(font(8))
        self._dirty_lbl = QLabel()
        self._dirty_lbl.setFont(font(8))
        self._dirty_lbl.setStyleSheet("color:#E74C3C;")
        tb.addWidget(self._dirty_lbl)
        tb.addWidget(self._status_lbl)
        layout.addLayout(tb)

        # 筛选栏
        fb = QHBoxLayout()
        fb.setSpacing(2)
        for lbl, items, attr in [("年级:", ["全部", "初一级", "初二级", "初三级"], "_grade_cb")]:
            fb.addWidget(self._lbl(lbl))
            cb = QComboBox()
            cb.addItems(items)
            cb.setFont(font(8))
            cb.setFixedWidth(90)
            cb.currentTextChanged.connect(lambda _: self._on_grade_change())
            setattr(self, attr, cb)
            fb.addWidget(cb)

        fb.addWidget(self._lbl("班级:"))
        self._class_cb = QComboBox()
        self._class_cb.setFont(font(8))
        self._class_cb.setFixedWidth(100)
        self._class_cb.addItem("全部")
        self._class_cb.currentTextChanged.connect(lambda _: self._apply_filter())
        fb.addWidget(self._class_cb)

        fb.addWidget(self._lbl("状态:"))
        self._status_cb = QComboBox()
        self._status_cb.setFont(font(8))
        self._status_cb.setFixedWidth(80)
        self._status_cb.addItems(["全部", "在校", "休学", "退学", "转学", "毕业"])
        self._status_cb.currentTextChanged.connect(lambda _: self._apply_filter())
        fb.addWidget(self._status_cb)

        fb.addWidget(self._lbl("搜索:"))
        self._search = QLineEdit()
        self._search.setFont(font(8))
        self._search.setPlaceholderText("姓名/学籍号/身份证/电话")
        self._search.setFixedWidth(180)
        self._search.setClearButtonEnabled(True)
        self._search.textChanged.connect(self._on_search_changed)
        fb.addWidget(self._search)

        fb.addStretch()
        layout.addLayout(fb)

        # 表格 + 详情（垂直分割）
        splitter = QSplitter(Qt.Vertical)
        self._table = self._make_table()
        splitter.addWidget(self._table)

        self._detail_panel = QFrame()
        self._detail_panel.setFrameShape(QFrame.StyledPanel)
        self._detail_panel.setStyleSheet(
            "background: white; border: 1px solid #DDD; border-radius: 4px;"
        )
        self._detail_layout = QVBoxLayout(self._detail_panel)
        self._detail_layout.setContentsMargins(10, 6, 10, 6)

        self._detail_title = QLabel("点击学生查看详情")
        self._detail_title.setFont(font(10, True))
        self._detail_layout.addWidget(self._detail_title)

        # 照片 + 信息（水平分割）
        detail_splitter = QSplitter(Qt.Horizontal)

        photo_card = QFrame()
        photo_card.setFixedWidth(160)
        pl = QVBoxLayout(photo_card)
        self._photo_label = QLabel("无照片")
        self._photo_label.setFixedSize(140, 170)
        self._photo_label.setAlignment(Qt.AlignCenter)
        self._photo_label.setStyleSheet(
            "border:1px solid #DDD; border-radius:4px; background:#F5F5F5;"
        )
        pl.addWidget(self._photo_label, alignment=Qt.AlignHCenter | Qt.AlignTop)
        pl.addStretch()
        detail_splitter.addWidget(photo_card)

        self._detail_text = QTextEdit()
        self._detail_text.setReadOnly(True)
        self._detail_text.setFont(font(9))
        self._detail_text.setContextMenuPolicy(Qt.CustomContextMenu)
        self._detail_text.customContextMenuRequested.connect(self._detail_context_menu)
        detail_splitter.addWidget(self._detail_text)

        detail_splitter.setSizes([160, 400])
        detail_splitter.setCollapsible(0, False)
        detail_splitter.setCollapsible(1, False)

        self._detail_layout.addWidget(detail_splitter)
        splitter.addWidget(self._detail_panel)
        splitter.setSizes([400, 300])
        layout.addWidget(splitter)

    def _on_grade_change(self):
        self._class_cb.blockSignals(True)
        self._class_cb.clear()
        self._class_cb.addItem("全部")
        g = self._grade_cb.currentText()
        if g != "全部":
            prefix = {"初一级": "1", "初二级": "2", "初三级": "3"}.get(g, "")
            if prefix:
                classes = (
                    self.session.query(ClassModel)
                    .filter(ClassModel.name.like(f"{prefix}%"))
                    .order_by(ClassModel.name)
                    .all()
                )
                for c in classes:
                    self._class_cb.addItem(c.name)
        self._class_cb.blockSignals(False)
        self._apply_filter()

    def _on_search_changed(self):
        self._apply_filter()

    def _make_table(self):
        from edu_system.gui.views.student_model import StudentTableModel

        t = QTableView()
        t.setStyleSheet(
            """QTableView { font-size:9pt; border:1px solid #DDD; gridline-color:#EEE;
            background:white; alternate-background-color:#EBF5FB; }
            QHeaderView::section { background: {C["table_header_bg"]}; font-weight:bold; font-size:9pt;
            padding:4px; border:1px solid {C["table_header_border"]}; color:#2C3E50; }
            QTableView::item { padding:2px 5px; }
            QTableView::item:selected { background:#3498DB; color:white; }"""
        )
        t.setAlternatingRowColors(True)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setSelectionMode(QAbstractItemView.ExtendedSelection)
        t.verticalHeader().hide()
        t.horizontalHeader().setStretchLastSection(False)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        t.setFont(font(9))
        t.clicked.connect(self._on_cell_clicked)
        t.doubleClicked.connect(self._on_cell_dblclicked)
        t.setContextMenuPolicy(Qt.CustomContextMenu)
        t.customContextMenuRequested.connect(self._enhanced_context_menu)
        t.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        t.horizontalHeader().sectionResized.connect(self._save_column_widths)
        self._model = StudentTableModel(self)
        t.setModel(self._model)
        t.verticalScrollBar().valueChanged.connect(self._on_scroll)
        t.selectionModel().selectionChanged.connect(self._on_selection_changed)
        return t

    def _on_cell_clicked(self, idx):
        s = self._model.get_student(idx.row())
        if s:
            self._show_detail_for_student(s)

    def _on_cell_dblclicked(self, idx):
        if self._model.get_student(idx.row()):
            self._edit_full()

    def _lbl(self, t):
        lay = QLabel(t)
        lay.setFont(font(8))
        lay.setStyleSheet("color:#555;")
        return lay

    # ═══════════════════════════════════
    # 表格增强
    # ═══════════════════════════════════

    def _on_header_clicked(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        key = self._visible_columns[col]
        self._data.sort(key=lambda s: getattr(s, key, ""), reverse=not self._sort_asc)
        self._render_table()
        self._save_sort_prefs(self._sort_col, self._sort_asc)

    def _save_column_widths(self, logical_index, old_size, new_size):
        from edu_system.models import Setting

        widths = []
        for i in range(self._model.columnCount()):
            widths.append(str(self._table.columnWidth(i)))
        val = ",".join(widths)
        entry = self.session.query(Setting).filter_by(key="student_table_col_widths").first()
        if entry:
            entry.value = val
        else:
            self.session.add(Setting(key="student_table_col_widths", value=val))
        self.session.commit()

    def _load_column_widths(self):
        from edu_system.models import Setting

        val = self.session.query(Setting).filter_by(key="student_table_col_widths").first()
        if val and val.value:
            widths = val.value.split(",")
            for i, w in enumerate(widths):
                if i < self._model.columnCount():
                    try:
                        self._table.setColumnWidth(i, int(w))
                    except:
                        pass

    def _restore_column_widths(self):
        self._load_column_widths()

    def _apply_saved_sort(self):
        if self._sort_col >= 0 and self._sort_col < len(self._visible_columns):
            key = self._visible_columns[self._sort_col]
            self._data.sort(key=lambda s: getattr(s, key, ""), reverse=not self._sort_asc)
            self._render_table()
            self._table.horizontalHeader().setSortIndicator(
                self._sort_col, Qt.AscendingOrder if self._sort_asc else Qt.DescendingOrder
            )

    def _on_scroll(self):
        vbar = self._table.verticalScrollBar()
        if vbar.value() >= vbar.maximum() - 5:
            self._load_next_page()

    def _on_selection_changed(self):
        self._update_status_bar()
        self._update_dirty_indicator()

    def _setup_table_enhancements(self):
        pass

    # ═══════════════════════════════════
    # 选中 / 详情 / 编辑
    # ═══════════════════════════════════

    def _selected(self):
        rows = set()
        for idx in self._table.selectionModel().selectedRows():
            rows.add(idx.row())
        return [self._data[r] for r in sorted(rows) if 0 <= r < len(self._data)]

    def _show_detail_for_student(self, s):
        info = [
            f"【{s.name}】 {s.gender}  状态: {s.status}",
            f"班级: {s.class_name}  座号: {s.student_no}  考号: {s.exam_no}",
            f"学籍号: {s.student_code}  身份证: {s.id_card}",
            f"民族: {s.ethnicity}  政治面貌: {s.political_status}",
            f"电话: {s.phone}  地址: {s.address}",
            f"户籍: {s.hukou_addr}  走住读: {s.boarding}",
            f"入学年: {s.enroll_year}",
            f"监护人1: {s.guardian1_name} ({s.guardian1_relation}) {s.guardian1_phone}",
            f"监护人2: {s.guardian2_name} ({s.guardian2_relation}) {s.guardian2_phone}",
        ]
        self._detail_title.setText(f"{s.name} ({s.class_name})")
        self._detail_text.setText("\n".join(info))

        if hasattr(self, "_photo_label") and s.photo:
            pixmap = QPixmap()
            pixmap.loadFromData(QByteArray(s.photo))
            if not pixmap.isNull():
                pixmap = pixmap.scaled(140, 170, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self._photo_label.setPixmap(pixmap)
                self._photo_label.setText("")
            else:
                self._photo_label.setText("无法加载")
                self._photo_label.setPixmap(QPixmap())
        elif hasattr(self, "_photo_label"):
            self._photo_label.setText("无照片")
            self._photo_label.setPixmap(QPixmap())

    def _detail_context_menu(self, pos):
        menu = QMenu(self)
        menu.addAction(
            "复制全部", lambda: QApplication.clipboard().setText(self._detail_text.toPlainText())
        )
        menu.addAction(
            "复制选中",
            lambda: QApplication.clipboard().setText(self._detail_text.textCursor().selectedText()),
        )
        menu.exec_(self._detail_text.mapToGlobal(pos))

    # ═══════════════════════════════════
    # 批量操作
    # ═══════════════════════════════════

    def _batch_transfer(self):
        students = self._selected()
        if not students:
            QMessageBox.information(self, "提示", "请先选择要转班的学生")
            return
        names = [c.name for c in self.session.query(ClassModel).order_by(ClassModel.name).all()]
        target, ok = QInputDialog.getItem(
            self, "批量转班", f"将 {len(students)} 名学生转到:", names, 0, False
        )
        if not ok:
            return
        cls = self.session.query(ClassModel).filter_by(name=target).first()
        if not cls:
            return
        from edu_system.services.enrollment import EnrollmentService

        svc = EnrollmentService(self.session)
        for s in students:
            svc.transfer(s.id, cls.id, "批量转班")
        self.session.commit()
        self._load_all()
        QMessageBox.information(self, "完成", f"已将 {len(students)} 名学生转入 {target}")

    def _batch_change_status(self):
        students = self._selected()
        if not students:
            QMessageBox.information(self, "提示", "请先选择要修改状态的学生")
            return
        statuses = ["在校", "休学", "复学", "退学", "毕业"]
        status, ok = QInputDialog.getItem(
            self, "批量改状态", f"将 {len(students)} 名学生设为:", statuses, 0, False
        )
        if not ok:
            return
        from edu_system.services.enrollment import EnrollmentService

        svc = EnrollmentService(self.session)
        for s in students:
            svc.change_status(s.id, status, "批量改状态")
        self.session.commit()
        self._load_all()
        QMessageBox.information(self, "完成", f"已将 {len(students)} 名学生状态改为 {status}")

    def _export_selected(self):
        students = self._selected()
        if not students:
            QMessageBox.information(self, "提示", "请先选择要导出的学生")
            return
        from openpyxl import Workbook

        dlg = QDialog(self)
        dlg.setWindowTitle("选择导出列")
        dlg.setMinimumWidth(420)
        layout = QVBoxLayout(dlg)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        cw = QWidget()
        cl = QVBoxLayout(cw)
        cl.setSpacing(2)
        vars = {}
        for key, label in ALL_COLUMNS.items():
            cb = QCheckBox(label)
            cb.setChecked(key in self._visible_columns)
            cb.setFont(font(9))
            cl.addWidget(cb)
            vars[key] = cb
        scroll.setWidget(cw)
        layout.addWidget(scroll)
        btn_row = QHBoxLayout()
        all_btn = QPushButton("全选")
        all_btn.clicked.connect(lambda: [cb.setChecked(True) for cb in vars.values()])
        none_btn = QPushButton("全不选")
        none_btn.clicked.connect(lambda: [cb.setChecked(False) for cb in vars.values()])
        btn_row.addWidget(all_btn)
        btn_row.addWidget(none_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)
        ok = QPushButton("导出")
        ok.setStyleSheet(
            f"background:{C['accent_green']}; color:white; border:none; border-radius:3px; padding:5px 15px;"
        )
        cancel = QPushButton("取消")
        cancel.clicked.connect(dlg.reject)
        bl = QHBoxLayout()
        bl.addWidget(ok)
        bl.addWidget(cancel)
        layout.addLayout(bl)

        def do_export():
            selected = [k for k, cb in vars.items() if cb.isChecked()]
            if not selected:
                QMessageBox.warning(self, "提示", "至少选一列")
                return
            dlg.accept()
            path, _ = QFileDialog.getSaveFileName(
                self, "保存", "选中学生名单.xlsx", "Excel (*.xlsx)"
            )
            if not path:
                return
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("学生名单")
            ws.append([f"选中学生名单 -- 共 {len(students)} 人"])
            ws.append([])
            ws.append([ALL_COLUMNS[k] for k in selected])
            for s in students:
                row = []
                for key in selected:
                    val = s.class_name if key == "class_name" else getattr(s, key, "")
                    if key == "birth_date" and val:
                        val = str(val)[:10]
                    row.append(str(val) if val else "")
                ws.append(row)
            wb.save(path)
            QMessageBox.information(
                self, "完成", f"已导出 {len(students)} 名学生 × {len(selected)} 列"
            )

        ok.clicked.connect(do_export)
        dlg.exec_()

    # ═══════════════════════════════════
    # 增强右键菜单
    # ═══════════════════════════════════

    def _enhanced_context_menu(self, pos):
        idx = self._table.indexAt(pos)
        if not idx.isValid():
            return
        menu = QMenu(self)
        students = self._selected()
        single = len(students) == 1
        s = students[0] if single else None

        if single:
            menu.addAction("查看详情", self._show_detail_dialog)
            menu.addAction("编辑信息", self._edit_full)
            menu.addSeparator()
            menu.addAction("复制姓名", lambda: self._copy_field(s, "name"))
            menu.addAction("复制学籍号", lambda: self._copy_field(s, "student_code"))
            menu.addAction("复制身份证", lambda: self._copy_field(s, "id_card"))
            menu.addAction("复制电话", lambda: self._copy_field(s, "phone"))
            menu.addSeparator()
            menu.addAction("快速转班...", lambda: self._quick_transfer(s))
            menu.addAction("快速改状态...", lambda: self._quick_change_status(s))
            menu.addAction("查看变动记录", lambda: self._show_movements(s))
            menu.addSeparator()

        if len(students) > 1:
            menu.addAction(f"批量转班 ({len(students)}人)", self._batch_transfer)
            menu.addAction(f"批量改状态 ({len(students)}人)", self._batch_change_status)
            menu.addAction(f"导出选中 ({len(students)}人)", self._export_selected)
            menu.addSeparator()

        menu.addAction("删除", self._delete_students)
        menu.exec_(self._table.viewport().mapToGlobal(pos))

    def _copy_field(self, student, field):
        val = getattr(student, field, "") or ""
        QApplication.clipboard().setText(str(val))
        QMessageBox.information(self, "已复制", f"{field}: {val}")

    def _quick_transfer(self, student):
        names = [c.name for c in self.session.query(ClassModel).order_by(ClassModel.name).all()]
        target, ok = QInputDialog.getItem(self, "转班", f"将 {student.name} 转到:", names, 0, False)
        if not ok:
            return
        cls = self.session.query(ClassModel).filter_by(name=target).first()
        if cls:
            from edu_system.services.enrollment import EnrollmentService

            EnrollmentService(self.session).transfer(student.id, cls.id, "快速转班")
            self.session.commit()
            self._load_all()

    def _quick_change_status(self, student):
        statuses = ["在校", "休学", "复学", "退学", "毕业"]
        target, ok = QInputDialog.getItem(
            self, "改状态", f"将 {student.name} 设为:", statuses, 0, False
        )
        if not ok:
            return
        from edu_system.services.enrollment import EnrollmentService

        EnrollmentService(self.session).change_status(student.id, target, "快速改状态")
        self.session.commit()
        self._load_all()

    def _show_movements(self, student):
        from PyQt5.QtWidgets import (
            QDialog,
            QHeaderView,
            QTableWidget,
            QTableWidgetItem,
            QVBoxLayout,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle(f"变动记录 - {student.name}")
        dlg.resize(600, 400)
        layout = QVBoxLayout(dlg)
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["日期", "类型", "详情", "原因"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        from edu_system.models import StudentMovement

        moves = (
            self.session.query(StudentMovement)
            .filter_by(student_id=student.id)
            .order_by(StudentMovement.created_at.desc())
            .all()
        )
        table.setRowCount(len(moves))
        for i, m in enumerate(moves):
            table.setItem(i, 0, QTableWidgetItem(str(m.move_date or "")[:10]))
            table.setItem(i, 1, QTableWidgetItem(m.move_type))
            detail = f"{m.from_class_id or ''} → {m.to_class_id or ''}"
            table.setItem(i, 2, QTableWidgetItem(detail))
            table.setItem(i, 3, QTableWidgetItem(m.reason or ""))
        layout.addWidget(table)
        dlg.exec_()

    def _show_detail_dialog(self):
        rows = set()
        for item in self._table.selectedItems():
            rows.add(item.row())
        if not rows:
            return
        self._show_detail_for_student(self._data[list(rows)[0]])

    def _edit_full(self):
        students = self._selected()
        if not students:
            return
        s = students[0]
        dlg = StudentEditDialog(self, s, "编辑", self.session)
        if dlg.exec_():
            for k, v in dlg.get_data().items():
                setattr(s, k, v)
            if s.id:
                self._dirty_ids.add(s.id)
            self._apply_filter()
            self._update_dirty_indicator()

    # ═══════════════════════════════════
    # 删除 / 新增
    # ═══════════════════════════════════

    def _delete_students(self):
        students = self._selected()
        if not students:
            return
        names = ", ".join(f"{s.name}({s.class_name})" for s in students[:5])
        if len(students) > 5:
            names += f" 等{len(students)}人"
        if (
            QMessageBox.question(self, "确认删除", f"标记删除以下学生？\n{names}")
            != QMessageBox.Yes
        ):
            return
        for s in students:
            if s.id:
                self._deleted_ids.add(s.id)
            if not s.id:
                self._new_students = [ns for ns in self._new_students if ns is not s]
        self._apply_filter()
        self._update_dirty_indicator()

    def _add(self):
        dlg = StudentEditDialog(self, None, "新增", self.session)
        if dlg.exec_():
            data = dlg.get_data()
            classes = self.session.query(ClassModel).order_by(ClassModel.name).all()
            cls = self.session.query(ClassModel).filter_by(name=data.get("class_name", "")).first()
            if not cls:
                return
            s = Student(class_id=cls.id, student_no="新")
            for k, v in data.items():
                if k != "class_name":
                    setattr(s, k, v)
            self._new_students.append(s)
            self._apply_filter()
            self._update_dirty_indicator()

    # ═══════════════════════════════════
    # 学生编辑对话框
    # ═══════════════════════════════════

    def _import(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入学生名单", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        photo_zip_path = None
        zip_path, _ = QFileDialog.getOpenFileName(self, "选择照片压缩包（可选）", "", "ZIP (*.zip)")
        if zip_path:
            photo_zip_path = zip_path

        from edu_system.services.importer import ImportService

        svc = ImportService(self.session)
        result = svc.import_students_from_excel(path, photo_zip_path=photo_zip_path)
        QMessageBox.information(
            self,
            "导入完成",
            f"成功: {result.succeeded}, 跳过: {result.skipped}, 冲突: {len(result.conflicts)}, 错误: {len(result.errors)}",
        )
        self._load_first_page()

    def _import_supplement(self):
        path, _ = QFileDialog.getOpenFileName(self, "导入补充信息", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        from edu_system.services.importer import ImportService

        svc = ImportService(self.session)
        result = svc.import_supplement(path)
        QMessageBox.information(
            self,
            "补充完成",
            f"更新: {result.succeeded}, 跳过: {result.skipped}, 错误: {len(result.errors)}",
        )
        self._load_first_page()

    def _export(self):
        from openpyxl import Workbook

        path, _ = QFileDialog.getSaveFileName(self, "保存", "学生名册.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("学生名册")
        ws.append([ALL_COLUMNS[k] for k in self._visible_columns])
        for s in self._all_data:
            if s.status == "在校" and s.id not in self._deleted_ids:
                row = []
                for key in self._visible_columns:
                    val = s.class_name if key == "class_name" else getattr(s, key, "")
                    if key == "birth_date" and val:
                        val = str(val)[:10]
                    row.append(str(val) if val else "")
                ws.append(row)
        wb.save(path)
        QMessageBox.information(self, "完成", f"已导出 {len(self._all_data)} 名学生")

    def _column_selector(self):
        dlg = ColumnSelectorDialog(self, ALL_COLUMNS, self._visible_columns, "选择显示列")
        if dlg.exec_():
            self._visible_columns = dlg.get_selected()
            self._save_column_prefs()
            self._render_table()

    def _generate_exam_nos(self):
        from sqlalchemy import text

        count = 0
        enroll_years = self.session.execute(
            text(
                "SELECT DISTINCT enroll_year FROM students WHERE status='在校' AND enroll_year > 0 ORDER BY enroll_year"
            )
        ).fetchall()
        for (enroll_year,) in enroll_years:
            prefix = str(enroll_year)[-2:]
            students = self.session.execute(
                text(
                    "SELECT s.id FROM students s JOIN classes c ON s.class_id=c.id "
                    "WHERE s.enroll_year=:ey AND s.status='在校' "
                    "ORDER BY c.name, CAST(s.student_no AS INTEGER)"
                ),
                {"ey": enroll_year},
            ).fetchall()
            for i, row in enumerate(students, 1):
                exam_no = f"{prefix}{i:04d}"
                self.session.execute(
                    text("UPDATE students SET exam_no=:no WHERE id=:id"),
                    {"no": exam_no, "id": row[0]},
                )
                count += 1
        self.session.commit()
        self._load_all()
        QMessageBox.information(self, "完成", f"已生成 {count} 个考号（按入学年份分组）")

    def _resort_seats(self):
        if (
            QMessageBox.question(self, "确认", "将各班学生按姓名拼音排序重新生成座号？")
            != QMessageBox.Yes
        ):
            return
        from pypinyin import Style, lazy_pinyin

        SURNAME = {
            "曾": "zeng",
            "区": "ou",
            "解": "xie",
            "单": "shan",
            "朴": "piao",
            "仇": "qiu",
            "查": "zha",
            "盖": "ge",
            "翟": "zhai",
            "乐": "yue",
            "卜": "bu3",
        }

        def py_key(name):
            f = name[0]
            if f in SURNAME:
                return SURNAME[f] + "".join(lazy_pinyin(name[1:], style=Style.TONE3))
            return "".join(lazy_pinyin(name, style=Style.TONE3))

        from sqlalchemy import text

        classes = self.session.execute(text("SELECT id FROM classes ORDER BY name")).fetchall()
        total = 0
        for (cls_id,) in classes:
            students = self.session.execute(
                text("SELECT id, name FROM students WHERE class_id=:cid AND status='在校'"),
                {"cid": cls_id},
            ).fetchall()
            students.sort(key=lambda r: py_key(r[1]))
            for i, stu_row in enumerate(students, 1):
                self.session.execute(
                    text("UPDATE students SET student_no=:no WHERE id=:id"),
                    {"no": str(i), "id": stu_row[0]},
                )
                total += 1
        self.session.commit()
        self._load_all()
        QMessageBox.information(self, "完成", f"已重排 {total} 名学生座号")

    def _print_roster(self):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side

        from edu_system.models import Semester

        cur = self.session.query(Semester).filter_by(is_active=True).first()
        sy = cur.year_start if cur else 2025
        sl = cur.semester if cur else "第二学期"
        wb = Workbook()
        wb.remove(wb.active)
        ft = Font(name="宋体", size=11, bold=True)
        fd = Font(name="宋体", size=10, bold=True)
        ac = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        medium = Side(style="medium")
        for cl in self.session.query(ClassModel).order_by(ClassModel.name).all():
            students = [
                s
                for s in self._all_data
                if s.class_name == cl.name and s.status == "在校" and s.id not in self._deleted_ids
            ]
            if not students:
                continue
            ws = wb.create_sheet(title=cl.name)
            gc = cl.name[0]
            ey = sy - (int(gc) - 1)
            cn = int(cl.name[1:])
            ws.merge_cells("A1:T1")
            ws["A1"].value = f"{self._school_name()}{ey}级{cn}班学生名单({sy}--{sy + 1}学年度{sl})"
            ws["A1"].font = ft
            ws["A1"].alignment = ac
            ws.row_dimensions[1].height = 36
            for c in range(1, 21):
                v = {1: "学号", 2: "姓名", 11: "学号", 12: "姓名"}.get(c, "")
                cell = ws.cell(row=2, column=c, value=v)
                cell.font = fd
                cell.alignment = ac
                cell.border = Border(
                    top=medium,
                    bottom=thin,
                    left=medium if c == 1 else thin,
                    right=medium if c == 20 else thin,
                )
            ws.row_dimensions[2].height = 22
            for i in range(25):
                r = 3 + i
                ws.row_dimensions[r].height = 22
                for side, off in [(0, 0), (1, 10)]:
                    idx = i if side == 0 else i + 25
                    if idx < len(students):
                        ws.cell(row=r, column=1 + off, value=students[idx].student_no).font = fd
                        ws.cell(row=r, column=2 + off, value=students[idx].name).font = fd
                    for cc in range(1 + off, 11 + off):
                        ws.cell(row=r, column=cc).alignment = ac
                        ws.cell(row=r, column=cc).border = Border(
                            top=thin, bottom=thin, left=thin, right=thin
                        )
        path, _ = QFileDialog.getSaveFileName(
            self,
            "保存名册",
            f"{self._school_name()}{sy}-{sy + 1}学年度{sl}学生名单.xlsx",
            "Excel (*.xlsx)",
        )
        if path:
            wb.save(path)
            QMessageBox.information(self, "完成", f"名册已保存:\n{path}")

    # ═══════════════════════════════════
    # 内存操作核心
    # ═══════════════════════════════════

    def _load_all(self):
        """全量加载所有学生到内存（零 SQL）"""
        if self._loading:
            return
        self._loading = True
        try:
            students = self._repo.cache.get_all_students_in_school()
            self._all_data = students
            self._total_count = len(students)

            sort_col, sort_asc = self._load_sort_prefs()
            if sort_col >= 0:
                self._sort_col = sort_col
                self._sort_asc = sort_asc
            self._apply_filter()
            self._update_dirty_indicator()
            self._load_column_widths()
        finally:
            self._loading = False

    def _load_first_page(self):
        """分页加载第一页（内存模式：直接切片）"""
        if self._loading:
            return
        self._loading = True
        try:
            # 内存模式：直接从缓存切片
            all_students = self._repo.cache.get_all_students_in_school()
            self._total_count = len(all_students)

            page_size = self._page_size
            self._all_data = all_students[:page_size]
            self._current_page = 1

            sort_col, sort_asc = self._load_sort_prefs()
            if sort_col >= 0:
                self._sort_col = sort_col
                self._sort_asc = sort_asc
            self._apply_filter()
            self._update_dirty_indicator()
            self._load_column_widths()
        finally:
            self._loading = False

    def _load_next_page(self):
        if len(self._all_data) >= self._total_count:
            return
        self._current_page += 1
        from edu_system.schemas import StudentPageRequest
        from edu_system.services.student import StudentRepository

        repo = StudentRepository(self.session)
        page_req = StudentPageRequest(page=self._current_page, page_size=self._page_size)
        page_result = repo.search_paginated(page_req)
        if page_result.ok and page_result.data:
            self._all_data.extend(page_result.data.items)
            self._apply_filter()

    def _apply_filter(self):
        data = list(self._all_data)
        data.extend(self._new_students)
        data = [s for s in data if s.id not in self._deleted_ids]

        grade = self._grade_cb.currentText() if self._grade_cb else "全部"
        cls_name = self._class_cb.currentText() if self._class_cb else "全部"
        status = self._status_cb.currentText() if self._status_cb else "全部"
        keyword = self._search.text().strip() if self._search else ""

        if grade != "全部":
            prefix = {"初一级": "1", "初二级": "2", "初三级": "3"}.get(grade, "")
            data = [s for s in data if s.class_name.startswith(prefix)]

        if cls_name and cls_name != "全部":
            data = [s for s in data if s.class_name == cls_name]

        if status and status != "全部":
            data = [s for s in data if s.status == status]

        if keyword:
            data = [
                s
                for s in data
                if keyword in s.name
                or keyword in s.student_code
                or keyword in s.id_card
                or keyword in s.phone
            ]

        self._data = data
        self._render_table()
        self._update_overview_in_memory()
        self._update_status_bar()

    def _render_table(self):
        self._model.set_data(self._data, self._visible_columns)
        if self._sort_col >= 0:
            self._table.horizontalHeader().setSortIndicator(
                self._sort_col, Qt.AscendingOrder if self._sort_asc else Qt.DescendingOrder
            )

    def _update_overview_in_memory(self):
        from collections import defaultdict

        by_grade = defaultdict(lambda: {"total": 0, "male": 0})
        for s in self._all_data:
            if s.status == "在校" and s.id not in self._deleted_ids:
                prefix = s.class_name[0] if s.class_name else "?"
                gn = {"1": "初一级", "2": "初二级", "3": "初三级"}.get(prefix, "未知")
                by_grade[gn]["total"] += 1
                if s.gender == "男":
                    by_grade[gn]["male"] += 1
        parts = []
        for g in ["初一级", "初二级", "初三级"]:
            if by_grade[g]["total"]:
                parts.append(f"{g}: {by_grade[g]['total']}人(男{by_grade[g]['male']})")
        self._overview.setText("  ".join(parts) if parts else "无数据")

    def _update_status_bar(self):
        total = len(self._all_data)
        filtered = len(self._data)
        selected = len(self._selected())
        self._status_lbl.setText(f"总计: {total}  |  筛选: {filtered}  |  选中: {selected}")

    def _update_dirty_indicator(self):
        dirty = len(self._dirty_ids) + len(self._deleted_ids) + len(self._new_students)
        self._dirty_lbl.setText(f"待保存: {dirty} 条" if dirty else "")

    def _save(self):
        if not (self._dirty_ids or self._deleted_ids or self._new_students):
            QMessageBox.information(self, "提示", "无待保存更改")
            return
        try:
            for s in self._new_students:
                self.session.add(s)
            for sid in self._dirty_ids:
                self.session.merge(self.session.query(Student).get(sid))
            for sid in self._deleted_ids:
                stu = self.session.query(Student).get(sid)
                if stu:
                    self.session.delete(stu)
            self.session.commit()
            self._dirty_ids.clear()
            self._deleted_ids.clear()
            self._new_students.clear()
            self._load_all()
            QMessageBox.information(self, "完成", "保存成功")
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self, "错误", f"保存失败: {e}")

    # ═══════════════════════════════════
    # 冲突解决对话框
    # ═══════════════════════════════════

    def _show_conflict_resolution_dialog(self, conflicts, parent):
        from PyQt5.QtWidgets import (
            QComboBox,
            QDialog,
            QHBoxLayout,
            QHeaderView,
            QLabel,
            QPushButton,
            QTableWidget,
            QTableWidgetItem,
            QVBoxLayout,
        )

        dlg = QDialog(parent)
        dlg.setWindowTitle("冲突解决 - 请选择处理方式")
        dlg.resize(900, 500)
        layout = QVBoxLayout(dlg)

        info = QLabel(
            f"检测到 {len(conflicts)} 行冲突数据。请为每行选择处理方式：\n"
            "• 覆盖：用新数据替换现有记录\n"
            "• 保留：保留数据库原有记录，跳过导入行\n"
            "• 手工编辑：在预览表中直接修改后再导入"
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(
            ["行号", "字段", "数据库值", "新值", "冲突类型", "处理方式"]
        )
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setRowCount(len(conflicts))

        for i, c in enumerate(conflicts):
            table.setItem(i, 0, QTableWidgetItem(str(c.row_index)))
            table.setItem(i, 1, QTableWidgetItem(c.name))
            table.setItem(i, 2, QTableWidgetItem(c.student_code))
            table.setItem(i, 3, QTableWidgetItem(c.class_name))
            table.setItem(i, 4, QTableWidgetItem(c.conflict_type))

            combo = QComboBox()
            combo.addItems(["保留原值", "覆盖现有", "手工编辑"])
            combo.setCurrentIndex(0)
            table.setCellWidget(i, 5, combo)

        layout.addWidget(table)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet(
            f"background:{C['accent_green']}; color:white; border:none; border-radius:3px; padding:5px 15px;"
        )
        cancel_btn = QPushButton("取消")
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        result = {}

        def on_ok():
            for i in range(len(conflicts)):
                combo = table.cellWidget(i, 5)
                text = combo.currentText()
                if text == "覆盖现有":
                    result[i] = "overwrite"
                elif text == "手工编辑":
                    result[i] = "manual"
                else:
                    result[i] = "keep"
            dlg.accept()

        def on_cancel():
            dlg.reject()

        ok_btn.clicked.connect(on_ok)
        cancel_btn.clicked.connect(on_cancel)

        if dlg.exec_() == dlg.Accepted:
            return result
        return None

    # ═══════════════════════════════════
    # 错误行导出
    # ═══════════════════════════════════

    def _export_error_rows(self, preview_table, headers, all_rows, data_start):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from PyQt5.QtWidgets import QFileDialog, QMessageBox

        error_rows = []
        for row in range(preview_table.rowCount()):
            status_item = preview_table.item(row, 7)
            if status_item and status_item.text().startswith("错误"):
                row_idx = row
                row_num = (
                    preview_table.item(row, 0).text()
                    if preview_table.item(row, 0)
                    else str(row + 1)
                )
                error_reason = status_item.toolTip() if status_item.toolTip() else "未知错误"

                orig_row_idx = data_start + row_idx
                if orig_row_idx < len(all_rows):
                    orig_data = all_rows[orig_row_idx]
                else:
                    orig_data = []

                error_rows.append(
                    {
                        "行号": row_num,
                        "错误原因": error_reason,
                        "原始数据": [str(v) if v is not None else "" for v in orig_data],
                    }
                )

        if not error_rows:
            QMessageBox.information(self, "提示", "没有错误行，无需导出")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "保存错误行", "导入错误行.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "错误行"

        error_headers = ["行号", "错误原因"] + [str(h) for h in headers]
        ws.append(error_headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(error_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, er in enumerate(error_rows, 2):
            ws.cell(row=i, column=1, value=er["行号"])
            ws.cell(row=i, column=2, value=er["错误原因"])
            for j, val in enumerate(er["原始数据"], 3):
                ws.cell(row=i, column=j, value=val)

        wb.save(path)
        QMessageBox.information(self, "完成", f"已导出 {len(error_rows)} 行错误数据到:\\n{path}")


# ═══════════════════════════════════
# 学生编辑对话框
# ═══════════════════════════════════
