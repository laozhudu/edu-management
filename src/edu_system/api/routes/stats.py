"""
统计 API 路由 — 学期统计查询 / 缓存管理 / 幂等重算触发
"""

import hashlib
import time
from datetime import datetime
from typing import Optional

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Body,
    Depends,
    HTTPException,
    Query,
    Request,
    Response,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from edu_system.api.deps import require_permission
from edu_system.core.permissions import Permission
from edu_system.database import get_active_semester, get_session
from edu_system.models import Semester, SemesterStatsCache
from edu_system.services.cache import bump_cache_version, cache_service, invalidate_stats_cache

router = APIRouter(prefix="/stats", tags=["统计数据"])


# ===== 依赖注入 =====


def get_db() -> Session:
    return get_session()


def get_current_semester() -> int:
    return get_active_semester()


# ===== 幂等键存储（内存，实际可用 Redis） =====
from collections import defaultdict

_idempotency_store = defaultdict(dict)
_IDEMPOTENCY_TTL = 3600  # 1 小时


def _check_idempotency(key: str) -> tuple[bool, dict | None]:
    """检查幂等键，返回 (是否已存在, 存储的结果)"""
    now = time.time()
    if key in _idempotency_store:
        entry = _idempotency_store[key]
        if now - entry["timestamp"] < _IDEMPOTENCY_TTL:
            return True, entry["result"]
        else:
            del _idempotency_store[key]
    return False, None


def _store_idempotency(key: str, result: dict):
    """存储幂等结果"""
    _idempotency_store[key] = {
        "result": result,
        "timestamp": time.time(),
    }


# ===== 统计重算触发接口 =====
_worker_instance: Optional["StatisticsWorker"] = None


def get_worker() -> "StatisticsWorker":
    global _worker_instance
    if _worker_instance is None:
        from edu_system.services.statistics import StatisticsWorker

        _worker_instance = StatisticsWorker()
    return _worker_instance


# ===== 统计数据查询接口 =====


@router.get("/semester/{semester_id}")
def get_semester_stats(
    semester_id: int,
    entity_type: str = Query("school", description="实体类型: school/grade/class/exam/subject"),
    entity_id: int = Query(0, description="实体ID，0表示学期汇总"),
    metric_keys: list[str] | None = Query(None, description="指标键列表，不传返回全部"),
    version: int | None = Query(None, description="缓存版本号，用于 304 判断"),
    request: Request = None,
    response: Response = None,
    db: Session = Depends(get_db),
):
    """
    获取学期统计数据
    支持 HTTP 304：前端传 version，服务端返回 304 如果版本一致
    """
    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(404, f"学期不存在: {semester_id}")

    current_version = cache_service.get_version()

    if version is not None and version == cache_service.get_version():
        return Response(status_code=304, headers={"ETag": f'W/"{cache_service.get_version()}"'})

    query = db.query(SemesterStatsCache).filter(
        SemesterStatsCache.semester_id == semester_id,
        SemesterStatsCache.entity_type == entity_type,
        SemesterStatsCache.entity_id == entity_id,
    )

    results = query.all()

    metrics = {}
    for r in results:
        metrics[r.metric_key] = {
            "value": r.metric_value,
            "version": r.version,
            "computed_at": r.computed_at.isoformat() if r.computed_at else None,
        }

    return {
        "semester_id": semester_id,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "version": cache_service.get_version(),
        "metrics": metrics,
    }


@router.get("/semester/{semester_id}/summary")
def get_semester_summary(
    semester_id: int,
    db: Session = Depends(get_db),
):
    """获取学期汇总统计（用于仪表盘概览）"""
    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(404, f"学期不存在: {semester_id}")

    from edu_system.services.cache import cache_service

    metrics = {}
    metric_keys = [
        "student_count",
        "student_male",
        "student_female",
        "class_count",
        "teacher_count",
        "subject_count",
        "score_avg",
        "score_pass_rate",
        "exam_count",
    ]

    for key in metric_keys:
        cached = cache_service.get("school", 0, key, semester_id=semester_id)
        if cached:
            metrics[key] = cached["value"]

    return {
        "semester_id": semester_id,
        "semester_label": f"{semester.year_start}-{semester.year_start + 1} 第{semester.semester}学期",
        "status": semester.status.value if hasattr(semester.status, "value") else semester.status,
        "metrics": metrics,
    }


# ===== 缓存管理接口 =====


@router.get("/cache/stats")
def get_cache_statistics():
    """获取缓存统计信息"""
    from edu_system.services.cache import cache_service

    return cache_service.get_stats()


@router.post("/cache/bump-version")
def bump_cache_version_api():
    """手动递增缓存版本（使所有缓存失效）"""
    new_version = bump_cache_version()
    return {"message": "缓存版本已递增", "new_version": cache_service.get_version()}


@router.post("/cache/invalidate")
def invalidate_cache(
    semester_id: int | None = None,
    tags: list[str] | None = None,
):
    """失效统计缓存"""
    sem_id = semester_id or get_active_semester()
    invalidate_stats_cache(semester_id=sem_id, tags=tags)
    return {"message": "缓存已失效", "semester_id": semester_id, "tags": tags}


@router.get("/cache/version")
def get_cache_version():
    """获取当前缓存版本"""
    from edu_system.services.cache import cache_service

    return {"version": cache_service.get_version()}


@router.get("/cache")
def get_cache(
    request: Request,
    response: Response,
):
    """获取缓存统计（支持 ETag/304）"""
    from edu_system.services.cache import cache_service

    stats = cache_service.get_stats()
    etag = f'W/"{cache_service.get_version()}"'

    if_none_match = request.headers.get("If-None-Match")
    if if_none_match and if_none_match == etag:
        return Response(status_code=304, headers={"ETag": etag})

    response.headers["ETag"] = etag
    return stats


@router.post("/recompute", status_code=200)
def trigger_recompute_idempotent(
    request: Request,
    background_tasks: BackgroundTasks,
    body: dict = Body(default={}),
    mode: str = "full",
    current_user: dict = Depends(require_permission(Permission.SYSTEM_ADMIN)),
):
    """
    幂等重算触发（M5-B4）

    - Idempotency-Key 头：相同键返回缓存结果（1 小时内）
    - mode=full: 全量重算
    - mode=incremental: 增量重算（需 dirty_entities）
    - 返回: {"status": "started"|"completed", "message": "...", "task_id": "..."}
    """
    idempotency_key = request.headers.get("Idempotency-Key")
    if not idempotency_key:
        raise HTTPException(
            status_code=400,
            detail="缺少 Idempotency-Key 头，请提供幂等键以防止重复提交",
        )

    exists, cached_result = _check_idempotency(idempotency_key)
    if exists and cached_result is not None:
        cached = dict(cached_result)
        cached["status"] = "completed"
        cached["message"] = "幂等键已存在，返回缓存结果"
        cached["task_id"] = idempotency_key
        return cached

    task_id = hashlib.sha256(f"{idempotency_key}:{time.time()}".encode()).hexdigest()[:16]

    worker = get_worker()

    dirty_entities = body.get("dirty_entities") if body else None

    if mode == "full":
        worker.start_full(
            progress_cb=lambda p, m: None,
            finished_cb=lambda m: invalidate_stats_cache(),
            error_cb=lambda e: print(f"[重算错误] {e}"),
        )
        result = {"message": "全量重算已在后台启动", "task": "full_recompute", "task_id": task_id}
    elif mode == "incremental":
        if not dirty_entities:
            raise HTTPException(status_code=400, detail="incremental 模式需要 dirty_entities")
        worker.start_incremental(
            dirty_entities=dirty_entities,
            progress_cb=lambda p, m: None,
            finished_cb=lambda m: invalidate_stats_cache(),
            error_cb=lambda e: print(f"错误: {e}"),
        )
        result = {"message": "增量重算已启动", "count": len(dirty_entities), "task_id": task_id}
    else:
        raise HTTPException(status_code=400, detail="mode 必须是 'full' 或 'incremental'")

    result["status"] = "started"
    _store_idempotency(idempotency_key, result)

    return {"status": "started", **result}


@router.post("/recompute/full")
def trigger_full_recompute(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_permission(Permission.SYSTEM_ADMIN)),
):
    """触发全量统计重算（后台任务）"""
    worker = get_worker()

    def on_finished(message: str):
        invalidate_stats_cache()

    def on_error(error: str):
        print(f"[重算错误] {error}")

    worker.start_full(
        progress_cb=lambda p, m: None,
        finished_cb=lambda m: invalidate_stats_cache(),
        error_cb=lambda e: print(f"错误: {e}"),
    )

    return {"message": "全量重算已在后台启动", "task": "full_recompute"}


@router.post("/recompute/incremental")
def trigger_incremental_recompute(
    dirty_entities: list[dict],
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_permission(Permission.SYSTEM_ADMIN)),
):
    """
    触发增量重算
    dirty_entities 格式: [{"entity_type": "class", "entity_id": 1}, {"entity_type": "exam", "entity_id": 5, "exam_id": 3}]
    """
    worker = get_worker()

    worker.start_incremental(
        dirty_entities=dirty_entities,
        progress_cb=lambda p, m: None,
        finished_cb=lambda m: invalidate_stats_cache(),
        error_cb=lambda e: print(f"错误: {e}"),
    )

    return {"message": "增量重算已启动", "count": len(dirty_entities)}


@router.post("/recompute/cancel")
def cancel_recompute():
    """取消正在进行的重算任务"""
    worker = get_worker()
    worker.cancel()
    return {"message": "已发送取消信号"}


@router.get("/recompute/worker/status")
def get_worker_status():
    """获取 Worker 状态"""
    worker = get_worker()
    return {
        "running": (
            worker._thread is not None and worker._thread.isRunning() if worker._thread else False
        ),
        "mode": worker._mode,
        "cancelled": worker._cancelled,
    }


# ===== 统计数据导出 =====


@router.get("/export/excel")
def export_stats_excel(
    semester_id: int,
    entity_type: str = Query("school"),
    entity_id: int = Query(0),
    db: Session = Depends(get_db),
):
    """导出统计数据为 Excel"""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    semester = db.query(Semester).filter(Semester.id == semester_id).first()
    if not semester:
        raise HTTPException(404, "学期不存在")

    wb = Workbook()
    ws = wb.active
    ws.title = "统计数据"

    headers = ["实体类型", "实体ID", "指标键", "指标值", "版本", "计算时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    query = db.query(SemesterStatsCache).filter(SemesterStatsCache.semester_id == semester_id)

    row = 2
    for stat in (
        db.query(SemesterStatsCache).filter(SemesterStatsCache.semester_id == semester_id).all()
    ):
        ws.cell(row=row, column=1, value=stat.entity_type)
        ws.cell(row=row, column=2, value=stat.entity_id)
        ws.cell(row=row, column=3, value=stat.metric_key)
        ws.cell(row=row, column=4, value=stat.metric_value)
        ws.cell(row=row, column=5, value=stat.version)
        ws.cell(
            row=row,
            column=6,
            value=stat.computed_at.strftime("%Y-%m-%d %H:%M:%S") if stat.computed_at else "",
        )
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stats_{semester_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/health")
def health_check():
    """健康检查"""
    from edu_system.services.cache import cache_service

    return {
        "status": "healthy",
        "cache_version": cache_service.get_version(),
        "cache_stats": cache_service.get_stats(),
        "timestamp": datetime.now().isoformat(),
    }
