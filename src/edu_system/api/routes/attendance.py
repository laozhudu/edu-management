"""
考勤 API 路由
"""

from datetime import date, datetime, time

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    WebSocket,
    WebSocketDisconnect,
    status,
)
from pydantic import BaseModel
from sqlalchemy import desc
from sqlalchemy.orm import Session

from edu_system.api.deps import get_current_user, get_db, require_permission
from edu_system.core.permissions import Permission
from edu_system.models import (
    Attendance,
    AttendanceStatus,
    AttendanceType,
    CheckMethod,
    Class,
    LeaveApplication,
    Student,
    User,
)

router = APIRouter(prefix="/attendance", tags=["考勤管理"])


# ===== Pydantic 模型 =====


class AttendanceCreate(BaseModel):
    student_id: int
    attendance_type: str = "morning"  # morning/noon/afternoon/evening/custom
    check_time: datetime | None = None
    scheduled_time: datetime | None = None
    check_method: str | None = None  # gps/bluetooth/face/qrcode/manual
    latitude: float | None = None
    longitude: float | None = None
    face_verified: bool = False
    device_info: str | None = None
    remark: str | None = None


class AttendanceUpdate(BaseModel):
    status: str | None = None
    check_time: datetime | None = None
    remark: str | None = None


class AttendanceResponse(BaseModel):
    id: int
    semester_id: int
    student_id: int
    class_id: int
    date: date
    attendance_type: str
    check_time: datetime | None
    scheduled_time: datetime | None
    status: str
    check_method: str | None
    latitude: float | None
    longitude: float | None
    face_verified: bool
    device_info: str | None
    remark: str | None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class AttendanceListResponse(BaseModel):
    items: list[AttendanceResponse]
    total: int
    page: int
    page_size: int


class LeaveApplicationCreate(BaseModel):
    student_id: int
    leave_type: str  # sick/personal/official/other
    start_date: date
    end_date: date
    start_time: time | None = None
    end_time: time | None = None
    reason: str
    attachments: list[str] | None = []


class LeaveApplicationUpdate(BaseModel):
    leave_type: str | None = None
    start_date: date | None = None
    end_date: date | None = None
    start_time: time | None = None
    end_time: time | None = None
    reason: str | None = None
    attachments: list[str] | None = None


class LeaveApplicationResponse(BaseModel):
    id: int
    semester_id: int
    student_id: int
    class_id: int
    leave_type: str
    start_date: date
    end_date: date
    start_time: time | None
    end_time: time | None
    reason: str
    attachments: list[str] | None
    status: str
    approved_by: int | None
    approved_at: datetime | None
    reject_reason: str | None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class LeaveApplicationListResponse(BaseModel):
    items: list[LeaveApplicationResponse]
    total: int
    page: int
    page_size: int


class LeaveApprovalRequest(BaseModel):
    approved: bool
    comment: str | None = None


class StatsRequest(BaseModel):
    scope: str = "class"  # student/class/grade/school
    student_id: int | None = None
    class_id: int | None = None
    grade_id: int | None = None
    date_from: date | None = None
    date_to: date | None = None


# ===== API 端点 =====


@router.post("/checkin", response_model=AttendanceResponse, status_code=status.HTTP_201_CREATED)
def checkin(
    checkin_data: AttendanceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.ATTENDANCE_ENTRY)),
):
    """打卡（学生/教师）

    自动判断状态：正常/迟到/早退/旷课
    支持 GPS/蓝牙/人脸/二维码/手工补录
    """
    # 验证学生
    student = db.query(Student).filter(Student.id == checkin_data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    # 获取活跃学期
    from edu_system.database import get_active_semester

    semester_id = get_active_semester()
    if not semester_id:
        raise HTTPException(status_code=400, detail="无活跃学期")

    # 确定考勤日期和类型
    check_time = checkin_data.check_time or datetime.now()
    attendance_date = check_time.date()
    attendance_type = checkin_data.attendance_type

    # 获取应打卡时间（简化：固定时间表）
    scheduled_times = {
        "morning": time(7, 30),
        "noon": time(12, 0),
        "afternoon": time(14, 0),
        "evening": time(19, 0),
    }
    scheduled_time = datetime.combine(
        attendance_date, scheduled_times.get(attendance_type, time(8, 0))
    )

    # 自动判断状态
    if checkin_data.face_verified or checkin_data.check_method == "face":
        status = AttendanceStatus.present
    else:
        diff_minutes = (check_time - scheduled_time).total_seconds() / 60
        if diff_minutes <= 0:
            status = AttendanceStatus.present
        elif diff_minutes <= 15:
            status = AttendanceStatus.late
        elif diff_minutes <= 60:
            status = (
                AttendanceStatus.early_leave
                if attendance_type in ["afternoon", "evening"]
                else AttendanceStatus.late
            )
        else:
            status = AttendanceStatus.absent

    # 检查是否已存在
    existing = (
        db.query(Attendance)
        .filter(
            Attendance.semester_id == semester_id,
            Attendance.student_id == checkin_data.student_id,
            Attendance.date == attendance_date,
            Attendance.attendance_type == AttendanceType(attendance_type),
        )
        .first()
    )

    if existing:
        # 更新现有记录（补卡/重复打卡）
        existing.check_time = check_time
        existing.scheduled_time = scheduled_time
        existing.status = status
        existing.check_method = (
            CheckMethod(checkin_data.check_method) if checkin_data.check_method else None
        )
        existing.latitude = checkin_data.latitude
        existing.longitude = checkin_data.longitude
        existing.face_verified = checkin_data.face_verified
        existing.device_info = checkin_data.device_info
        existing.remark = checkin_data.remark
        existing.updated_at = datetime.now()
        db.commit()
        db.refresh(existing)
        return existing

    # 创建新记录
    attendance = Attendance(
        semester_id=semester_id,
        student_id=checkin_data.student_id,
        class_id=student.class_id,
        date=attendance_date,
        attendance_type=AttendanceType(attendance_type),
        check_time=check_time,
        scheduled_time=scheduled_time,
        status=status,
        check_method=CheckMethod(checkin_data.check_method) if checkin_data.check_method else None,
        latitude=checkin_data.latitude,
        longitude=checkin_data.longitude,
        face_verified=checkin_data.face_verified,
        device_info=checkin_data.device_info,
        remark=checkin_data.remark,
    )
    db.add(attendance)
    db.commit()
    db.refresh(attendance)
    _push_checkin_event(attendance, status, student.name)
    return attendance


def _push_checkin_event(attendance, status, student_name: str):
    """打卡后推送实时事件给订阅者（同步路由内调用，离线自动入队）"""
    try:
        import asyncio

        from edu_system.services.attendance_notifier import get_notifier

        notifier = get_notifier()
        event = {
            "type": "attendance.checkin",
            "data": {
                "student_id": attendance.student_id,
                "student_name": student_name,
                "date": attendance.date.isoformat() if attendance.date else None,
                "attendance_type": attendance.attendance_type,
                "status": str(status),
                "check_time": attendance.check_time.isoformat()
                if attendance.check_time
                else None,
            },
        }
        # 尽力推送：有运行事件循环用 create_task，否则同步等待
        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            asyncio.run(notifier.notify(event))
        else:
            loop.create_task(notifier.notify(event))
    except Exception:
        # 推送失败不影响主流程
        pass


@router.get("", response_model=AttendanceListResponse)
def list_attendance(
    student_id: int | None = Query(None),
    class_id: int | None = Query(None),
    attendance_type: str | None = Query(None),
    status: str | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """考勤记录列表查询（分页/筛选）"""
    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    query = db.query(Attendance).filter(Attendance.semester_id == semester_id)

    # 权限过滤
    if current_user.role.name == "teacher":
        # 教师只能看自己任课班级
        # 简化：通过 ClassSubject 关联
        pass
    elif current_user.role.name == "student":
        # 学生只能看自己
        query = query.filter(Attendance.student_id == current_user.id)

    if student_id:
        query = query.filter(Attendance.student_id == student_id)
    if class_id:
        query = query.filter(Attendance.class_id == class_id)
    if attendance_type:
        query = query.filter(Attendance.attendance_type == AttendanceType(attendance_type))
    if status:
        query = query.filter(Attendance.status == AttendanceStatus(status))
    if date_from:
        query = query.filter(Attendance.date >= date_from)
    if date_to:
        query = query.filter(Attendance.date <= date_to)

    total = query.count()
    records = (
        query.order_by(desc(Attendance.date), desc(Attendance.check_time))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return AttendanceListResponse(
        items=records,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/leave", response_model=LeaveApplicationListResponse)
def list_leave_applications(
    student_id: int | None = Query(None),
    class_id: int | None = Query(None),
    status: str | None = Query(None),
    leave_type: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """请假申请列表查询"""
    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    query = db.query(LeaveApplication).filter(LeaveApplication.semester_id == semester_id)

    # 权限过滤
    if current_user.role.name == "student":
        query = query.filter(LeaveApplication.student_id == current_user.id)
    elif current_user.role.name == "teacher":
        # 班主任看本班
        pass

    if student_id:
        query = query.filter(LeaveApplication.student_id == student_id)
    if class_id:
        query = query.filter(LeaveApplication.class_id == class_id)
    if status:
        query = query.filter(LeaveApplication.status == status)
    if leave_type:
        query = query.filter(LeaveApplication.leave_type == leave_type)

    total = query.count()
    leaves = (
        query.order_by(desc(LeaveApplication.created_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return LeaveApplicationListResponse(
        items=leaves,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("/leave", response_model=LeaveApplicationResponse, status_code=status.HTTP_201_CREATED)
def create_leave_application(
    leave_data: LeaveApplicationCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.ATTENDANCE_ENTRY)),
):
    """提交请假申请"""
    student = db.query(Student).filter(Student.id == leave_data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    # 验证日期
    if leave_data.end_date < leave_data.start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")

    import json

    leave = LeaveApplication(
        semester_id=semester_id,
        student_id=leave_data.student_id,
        class_id=student.class_id,
        leave_type=leave_data.leave_type,
        start_date=leave_data.start_date,
        end_date=leave_data.end_date,
        start_time=leave_data.start_time,
        end_time=leave_data.end_time,
        reason=leave_data.reason,
        attachments=json.dumps(leave_data.attachments) if leave_data.attachments else None,
        status="pending",
    )
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return leave


@router.put("/leave/{leave_id}/approve", response_model=LeaveApplicationResponse)
def approve_leave_application(
    leave_id: int,
    approval: LeaveApprovalRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.ATTENDANCE_APPROVE)),
):
    """审批请假申请（通过/驳回）"""
    leave = db.query(LeaveApplication).filter(LeaveApplication.id == leave_id).first()
    if not leave:
        raise HTTPException(status_code=404, detail="请假申请不存在")

    if leave.status != "pending":
        raise HTTPException(status_code=400, detail="申请已处理")

    if approval.approved:
        leave.status = "approved"
    else:
        leave.status = "rejected"
        leave.reject_reason = approval.comment

    leave.approved_by = current_user.id
    leave.approved_at = datetime.now()
    db.commit()
    db.refresh(leave)
    return leave


@router.get("/stats")
def attendance_stats(
    scope: str = Query("class"),
    student_id: int | None = Query(None),
    class_id: int | None = Query(None),
    grade_id: int | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.ATTENDANCE_STATS)),
):
    """考勤统计（个人/班级/年级/全校）"""
    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    query = db.query(Attendance).filter(Attendance.semester_id == semester_id)

    if date_from:
        query = query.filter(Attendance.date >= date_from)
    if date_to:
        query = query.filter(Attendance.date <= date_to)

    if scope == "student" and student_id:
        query = query.filter(Attendance.student_id == student_id)
    elif scope == "class" and class_id:
        query = query.filter(Attendance.class_id == class_id)
    elif scope == "grade" and grade_id:
        # 通过班级关联年级
        from edu_system.models import Class as ClassModel

        class_ids = db.query(ClassModel.id).filter(ClassModel.grade_id == grade_id).subquery()
        query = query.filter(Attendance.class_id.in_(class_ids))
    elif scope == "school":
        pass  # 全校

    records = query.all()
    total = len(records)

    if total == 0:
        return {
            "total": 0,
            "present": 0,
            "late": 0,
            "early_leave": 0,
            "absent": 0,
            "leave": 0,
            "rate": 0,
        }

    present = sum(1 for r in records if r.status == AttendanceStatus.present)
    late = sum(1 for r in records if r.status == AttendanceStatus.late)
    early_leave = sum(1 for r in records if r.status == AttendanceStatus.early_leave)
    absent = sum(1 for r in records if r.status == AttendanceStatus.absent)
    leave = sum(1 for r in records if r.status == AttendanceStatus.leave)

    return {
        "total": total,
        "present": present,
        "late": late,
        "early_leave": early_leave,
        "absent": absent,
        "leave": leave,
        "present_rate": round(present / total * 100, 2),
        "late_rate": round(late / total * 100, 2),
        "absent_rate": round(absent / total * 100, 2),
    }


@router.get("/export")
def export_attendance(
    class_id: int | None = Query(None),
    student_id: int | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_REPORT)),
):
    """导出考勤 Excel"""
    import io

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    query = (
        db.query(Attendance).join(Student).join(Class).filter(Attendance.semester_id == semester_id)
    )

    if class_id:
        query = query.filter(Attendance.class_id == class_id)
    if student_id:
        query = query.filter(Attendance.student_id == student_id)
    if date_from:
        query = query.filter(Attendance.date >= date_from)
    if date_to:
        query = query.filter(Attendance.date <= date_to)

    records = query.order_by(
        Attendance.date, Class.name, Student.name, Attendance.attendance_type
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤导出"

    headers = [
        "日期",
        "考勤类型",
        "班级",
        "学生",
        "学号",
        "应打卡时间",
        "实打卡时间",
        "状态",
        "打卡方式",
        "定位",
        "备注",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.date.strftime("%Y-%m-%d"))
        ws.cell(
            row=row_idx,
            column=2,
            value=record.attendance_type.value if record.attendance_type else "",
        )
        ws.cell(row=row_idx, column=3, value=record.class_.name if record.class_ else "")
        ws.cell(row=row_idx, column=4, value=record.student.name)
        ws.cell(row=row_idx, column=5, value=record.student.student_code)
        ws.cell(
            row=row_idx,
            column=6,
            value=record.scheduled_time.strftime("%H:%M") if record.scheduled_time else "",
        )
        ws.cell(
            row=row_idx,
            column=7,
            value=record.check_time.strftime("%H:%M") if record.check_time else "",
        )
        ws.cell(row=row_idx, column=8, value=record.status.value if record.status else "")
        ws.cell(
            row=row_idx, column=9, value=record.check_method.value if record.check_method else ""
        )
        location = ""
        if record.latitude and record.longitude:
            location = f"{record.latitude},{record.longitude}"
        ws.cell(row=row_idx, column=10, value=location)
        ws.cell(row=row_idx, column=11, value=record.remark or "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{date_from or 'all'}_{date_to or ''}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/alerts")
def attendance_alerts(
    date: date | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.ATTENDANCE_VIEW)),
):
    """异常打卡告警（迟到/早退/旷课）"""
    from edu_system.database import get_active_semester

    semester_id = get_active_semester()

    if not date:
        date = datetime.now().date()

    query = (
        db.query(Attendance)
        .join(Student)
        .join(Class)
        .filter(
            Attendance.semester_id == semester_id,
            Attendance.date == date,
            Attendance.status.in_(
                [AttendanceStatus.late, AttendanceStatus.early_leave, AttendanceStatus.absent]
            ),
        )
    )

    alerts = query.order_by(Attendance.status, Class.name, Student.name).all()

    return {
        "date": date.isoformat(),
        "total": len(alerts),
        "late": [a for a in alerts if a.status == AttendanceStatus.late],
        "early_leave": [a for a in alerts if a.status == AttendanceStatus.early_leave],
        "absent": [a for a in alerts if a.status == AttendanceStatus.absent],
    }


@router.websocket("/ws")
async def attendance_ws(websocket: WebSocket):
    """考勤实时推送（WebSocket，M5-E2）

    连接后订阅考勤事件（checkin 时推送），断线重连自动补发离线期间事件。
    鉴权：query 参数 token（?token=xxx，空 token 拒绝）。
    """
    from edu_system.services.attendance_notifier import get_notifier

    token = websocket.query_params.get("token", "")
    if not token:
        await websocket.close(code=4401)
        return

    notifier = get_notifier()
    await websocket.accept()
    await notifier.register(websocket)
    try:
        while True:
            # 保持连接；客户端心跳/消息（目前仅用于维持）
            await websocket.receive_text()
    except WebSocketDisconnect:
        await notifier.unregister(websocket)
    except Exception:
        await notifier.unregister(websocket)


if __name__ == "__main__":
    pytest.main([__file__, "-x", "-v"])
