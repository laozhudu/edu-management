"""
成绩 API 路由
"""

from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel
from sqlalchemy import desc
from sqlalchemy.orm import Session

from edu_system.api.deps import get_current_user, get_db, require_permission
from edu_system.core.permissions import Permission
from edu_system.models import Class, Exam, Score, Student, Subject, User

router = APIRouter(prefix="/score", tags=["成绩管理"])


# ===== Pydantic 模型 =====


class ScoreCreate(BaseModel):
    exam_id: int
    student_id: int
    subject_id: int
    score: float
    is_makeup: bool = False


class ScoreUpdate(BaseModel):
    score: float | None = None
    is_makeup: bool | None = None


class ScoreResponse(BaseModel):
    id: int
    exam_id: int
    student_id: int
    subject_id: int
    score: float
    is_makeup: bool
    is_published: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class ScoreListResponse(BaseModel):
    items: list[ScoreResponse]
    total: int
    page: int
    page_size: int


class BatchScoreCreate(BaseModel):
    scores: list[ScoreCreate]


class PasteScoresRequest(BaseModel):
    """Excel 粘贴录入：TSV/CSV 文本

    每行: 学号<TAB>科目名<TAB>成绩[<TAB>是否补考(0/1)]
    首行可含表头（自动跳过：学号/科目/成绩 关键字）
    """

    exam_id: int
    text: str
    delimiter: str = "\t"  # 支持 \t 或 ,


class RankRequest(BaseModel):
    exam_id: int
    scope: str = "class"  # class/grade/school


class PublishRequest(BaseModel):
    exam_id: int
    published: bool


# ===== API 端点 =====


@router.get("", response_model=ScoreListResponse)
def list_scores(
    exam_id: int | None = Query(None),
    student_id: int | None = Query(None),
    subject_id: int | None = Query(None),
    class_id: int | None = Query(None),
    is_makeup: bool | None = Query(None),
    is_published: bool | None = Query(None),
    keyword: str | None = Query(None, description="关键字：学生姓名/学号/班级名模糊搜索"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """成绩列表查询（分页/筛选/排序）"""
    query = db.query(Score).join(Exam).join(Student).join(Subject)

    # 权限过滤
    if current_user.role.name == "teacher":
        # 教师只能看自己任课班级的成绩
        # 这里简化处理，实际需关联 ClassSubject 表
        pass

    if exam_id:
        query = query.filter(Score.exam_id == exam_id)
    if student_id:
        query = query.filter(Score.student_id == student_id)
    if subject_id:
        query = query.filter(Score.subject_id == subject_id)
    if is_makeup is not None:
        query = query.filter(Score.is_makeup == is_makeup)
    if is_published is not None:
        query = query.filter(Score.is_published == is_published)

    # 班级筛选（通过学生关联）
    if class_id:
        query = query.filter(Student.class_id == class_id)

    # 关键字模糊搜索（学生姓名/学号/班级名）
    if keyword:
        from sqlalchemy import or_

        kw = f"%{keyword.strip()}%"
        # 需 join Class（通过 Student.class_id）以支持班级名搜索
        query = query.outerjoin(Class, Student.class_id == Class.id)
        query = query.filter(
            or_(
                Student.name.like(kw),
                Student.student_no.like(kw),
                Class.name.like(kw),
            )
        )

    total = query.count()
    scores = (
        query.order_by(desc(Score.created_at)).offset((page - 1) * page_size).limit(page_size).all()
    )

    return ScoreListResponse(
        items=scores,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/export")
def export_scores(
    exam_id: int | None = Query(None),
    class_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_REPORT)),
):
    """导出成绩 Excel"""
    import io

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    query = db.query(Score).join(Exam).join(Student).join(Subject).join(Class)

    if exam_id:
        query = query.filter(Score.exam_id == exam_id)
    if class_id:
        query = query.filter(Student.class_id == class_id)

    scores = query.order_by(Exam.name, Class.name, Student.name, Subject.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导出"

    headers = ["考试", "班级", "学生", "学号", "科目", "分数", "补考", "发布状态", "录入时间"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, score in enumerate(scores, 2):
        ws.cell(row=row_idx, column=1, value=score.exam.name if score.exam else "")
        ws.cell(
            row=row_idx, column=2, value=score.student.class_.name if score.student.class_ else ""
        )
        ws.cell(row=row_idx, column=3, value=score.student.name)
        ws.cell(row=row_idx, column=4, value=score.student.student_code)
        ws.cell(row=row_idx, column=5, value=score.subject.name if score.subject else "")
        ws.cell(row=row_idx, column=6, value=score.score)
        ws.cell(row=row_idx, column=7, value="是" if score.is_makeup else "否")
        ws.cell(row=row_idx, column=8, value="已发布" if score.is_published else "未发布")
        ws.cell(row=row_idx, column=9, value=score.created_at.strftime("%Y-%m-%d %H:%M"))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"scores_{exam_id or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{score_id}", response_model=ScoreResponse)
def get_score(
    score_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取单条成绩"""
    score = db.query(Score).filter(Score.id == score_id).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩不存在")
    return score


@router.post("", response_model=ScoreResponse, status_code=status.HTTP_201_CREATED)
def create_score(
    score_data: ScoreCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """创建成绩"""
    # 验证关联对象存在
    exam = db.query(Exam).filter(Exam.id == score_data.exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    student = db.query(Student).filter(Student.id == score_data.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    subject = db.query(Subject).filter(Subject.id == score_data.subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="科目不存在")

    # 检查是否已存在
    existing = (
        db.query(Score)
        .filter(
            Score.exam_id == score_data.exam_id,
            Score.student_id == score_data.student_id,
            Score.subject_id == score_data.subject_id,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="该成绩已存在")

    # 检查数据锁
    from edu_system.services.locks import DataLockService

    lock = DataLockService.check_lock(db, "score", score_data.exam_id)
    if lock and lock.lock_level.value == "hard":
        raise HTTPException(status_code=403, detail="成绩已硬锁定，不可修改")

    score = Score(
        exam_id=score_data.exam_id,
        student_id=score_data.student_id,
        subject_id=score_data.subject_id,
        score=score_data.score,
        is_makeup=score_data.is_makeup,
        is_published=False,
    )
    db.add(score)
    db.commit()
    db.refresh(score)
    return score


@router.put("/{score_id}", response_model=ScoreResponse)
def update_score(
    score_id: int,
    score_data: ScoreUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """更新成绩"""
    score = db.query(Score).filter(Score.id == score_id).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩不存在")

    # 检查数据锁
    from edu_system.services.locks import DataLockService

    lock = DataLockService.check_lock(db, "score", score.exam_id)
    if lock and lock.lock_level.value == "hard":
        raise HTTPException(status_code=403, detail="成绩已硬锁定，不可修改")
    if lock and lock.lock_level.value == "soft" and current_user.role.name != "admin":
        raise HTTPException(status_code=403, detail="成绩已软锁定，需管理员确认")

    if score_data.score is not None:
        score.score = score_data.score
    if score_data.is_makeup is not None:
        score.is_makeup = score_data.is_makeup

    db.commit()
    db.refresh(score)
    return score


@router.delete("/{score_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_score(
    score_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """删除成绩（软删除：标记为未发布）"""
    score = db.query(Score).filter(Score.id == score_id).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩不存在")

    # 检查数据锁
    from edu_system.services.locks import DataLockService

    lock = DataLockService.check_lock(db, "score", score.exam_id)
    if lock and lock.lock_level.value == "hard":
        raise HTTPException(status_code=403, detail="成绩已硬锁定，不可删除")

    # 软删除：标记未发布，实际保留记录
    score.is_published = False
    db.commit()
    return None


@router.post("/batch", response_model=list[ScoreResponse], status_code=status.HTTP_201_CREATED)
def batch_create_scores(
    batch_data: BatchScoreCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """批量创建/更新成绩（Excel 导入）"""
    # 锁定检查：考试级硬锁/软锁拦截批量写入（E1 验收「锁定」）
    from edu_system.services.locks import DataLockService

    exam_ids = {s.exam_id for s in batch_data.scores}
    for eid in exam_ids:
        lock = DataLockService.check_lock(db, "score", eid)
        if lock and lock.lock_level.value == "hard":
            raise HTTPException(status_code=403, detail=f"考试 {eid} 成绩已硬锁定，不可修改")
        if lock and lock.lock_level.value == "soft" and current_user.role.name != "admin":
            raise HTTPException(status_code=403, detail=f"考试 {eid} 成绩已软锁定，需管理员确认")

    results = []
    errors = []

    for idx, score_data in enumerate(batch_data.scores):
        try:
            # 检查是否已存在
            existing = (
                db.query(Score)
                .filter(
                    Score.exam_id == score_data.exam_id,
                    Score.student_id == score_data.student_id,
                    Score.subject_id == score_data.subject_id,
                )
                .first()
            )

            if existing:
                # 更新现有
                existing.score = score_data.score
                existing.is_makeup = score_data.is_makeup
                score = existing
            else:
                # 创建新
                score = Score(
                    exam_id=score_data.exam_id,
                    student_id=score_data.student_id,
                    subject_id=score_data.subject_id,
                    score=score_data.score,
                    is_makeup=score_data.is_makeup,
                    is_published=False,
                )
                db.add(score)

            db.flush()
            results.append(score)
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    db.commit()

    for score in results:
        db.refresh(score)

    if errors:
        return {"created": len(results), "errors": errors}

    return results


@router.post("/paste")
def paste_scores(
    request: PasteScoresRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """Excel 粘贴录入：TSV/CSV 文本 → 批量创建/更新成绩（E1）

    每行: 学号<TAB>科目名<TAB>成绩[<TAB>补考(0/1)]
    首行含表头自动跳过（学号/科目/成绩关键字）
    返回: {created, updated, errors:[{row,error}]}
    """
    from edu_system.services.locks import DataLockService

    delimeter = request.delimiter.replace("\\t", "\t")
    lines = [ln.strip() for ln in request.text.strip().splitlines() if ln.strip()]
    if not lines:
        raise HTTPException(status_code=400, detail="粘贴内容为空")

    # 跳过表头（首行含关键字）
    head_kw = ("学号", "科目", "成绩", "姓名", "student", "subject", "score")
    if any(k.lower() in lines[0].lower() for k in head_kw):
        lines = lines[1:]

    results, errors, updated = [], [], 0
    for idx, line in enumerate(lines):
        parts = [p.strip() for p in line.split(delimeter)]
        if len(parts) < 3:
            errors.append({"row": idx + 2, "error": f"列数不足: {len(parts)}（需 学号/科目/成绩）"})
            continue
        student_no, subject_name, score_str = parts[0], parts[1], parts[2]
        is_makeup = parts[3] in ("1", "true", "是") if len(parts) > 3 else False

        student = db.query(Student).filter(Student.student_no == student_no).first()
        if not student:
            errors.append({"row": idx + 2, "error": f"学号不存在: {student_no}"})
            continue
        subject = db.query(Subject).filter(Subject.name == subject_name).first()
        if not subject:
            errors.append({"row": idx + 2, "error": f"科目不存在: {subject_name}"})
            continue
        try:
            score_val = float(score_str)
        except ValueError:
            errors.append({"row": idx + 2, "error": f"成绩非数字: {score_str}"})
            continue

        # 锁定检查（考试级）
        lock = DataLockService.check_lock(db, "score", request.exam_id)
        if lock and lock.lock_level.value == "hard":
            errors.append({"row": idx + 2, "error": "成绩已硬锁定，不可修改"})
            continue
        if lock and lock.lock_level.value == "soft" and current_user.role.name != "admin":
            errors.append({"row": idx + 2, "error": "成绩已软锁定，需管理员确认"})
            continue

        existing = (
            db.query(Score)
            .filter(
                Score.exam_id == request.exam_id,
                Score.student_id == student.id,
                Score.subject_id == subject.id,
            )
            .first()
        )
        if existing:
            existing.score = score_val
            existing.is_makeup = is_makeup
            updated += 1
            results.append(existing)
        else:
            score = Score(
                exam_id=request.exam_id,
                student_id=student.id,
                subject_id=subject.id,
                score=score_val,
                is_makeup=is_makeup,
                is_published=False,
            )
            db.add(score)
            results.append(score)

    db.commit()
    for s in results:
        db.refresh(s)
    return {
        "created": len(results) - updated,
        "updated": updated,
        "errors": errors,
        "total_rows": len(lines),
    }


@router.post("/rank")
def calculate_rank(
    request: RankRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """排名计算（班级/年级/学科）"""
    # 这里简化实现，实际应调用 StatisticsService
    from edu_system.services.statistics import StatisticsService

    stats_service = StatisticsService(db)
    result = stats_service.calculate_rank(
        exam_id=request.exam_id,
        scope=request.scope,
    )

    return {"message": "排名计算完成", "result": result}


@router.post("/publish")
def publish_scores(
    request: PublishRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission(Permission.SCORE_ENTRY)),
):
    """发布/取消发布成绩"""
    updated = (
        db.query(Score)
        .filter(Score.exam_id == request.exam_id)
        .update({"is_published": request.published})
    )
    db.commit()

    return {"message": f"已{'发布' if request.published else '取消发布'} {updated} 条成绩"}


if __name__ == "__main__":
    pytest.main([__file__, "-x", "-v"])
