"""
导入清洗服务 — 从外部文件到数据库的完整编排
"""

from dataclasses import dataclass
from enum import Enum

from sqlalchemy.orm import Session

from edu_system.core import Permission, require_permission
from edu_system.core.audit import manual_audit
from edu_system.models import Class as ClassModel
from edu_system.models import Grade, Student, Subject


class ConflictStrategy(Enum):
    """冲突解决策略"""

    SKIP = "skip"  # 跳过该行（保持原有）
    OVERWRITE = "overwrite"  # 覆盖现有记录
    KEEP_EXISTING = "keep"  # 保留现有，跳过导入行


@dataclass
class ConflictInfo:
    """冲突信息"""

    row_index: int
    name: str
    student_code: str
    class_name: str
    conflict_type: str  # "student_code" / "class_name" / "name_cross_class"
    existing_student: "Student"
    new_data: dict
    strategy: ConflictStrategy = ConflictStrategy.SKIP


class ImportResult:
    """导入结果"""

    def __init__(self):
        self.total = 0
        self.succeeded = 0
        self.skipped = 0
        self.conflicts: list[ConflictInfo] = []
        self.errors: list[str] = []
        self.checkpoint: dict = {}  # 断点续传数据

    @property
    def summary(self) -> str:
        lines = [
            f"总计: {self.total}  成功: {self.succeeded}  跳过: {self.skipped}  冲突: {len(self.conflicts)}  失败: {len(self.errors)}"
        ]
        for e in self.errors[:10]:
            lines.append(f"  ❌ {e}")
        if len(self.errors) > 10:
            lines.append(f"  ... 还有 {len(self.errors) - 10} 个错误")
        for c in self.conflicts[:5]:
            lines.append(f"  ⚠️ 冲突: {c.name} ({c.conflict_type}) - 策略: {c.strategy.value}")
        if len(self.conflicts) > 5:
            lines.append(f"  ... 还有 {len(self.conflicts) - 5} 个冲突")
        return "\n".join(lines)


class ImportService:
    def __init__(self, session: Session):
        self.session = session
        self._cancelled = False
        self._progress_callback = None
        self._checkpoint_interval = 50  # 每 50 行保存一次断点

    @require_permission(Permission.STUDENT_IMPORT)
    def import_students_from_excel(
        self,
        path: str,
        photo_zip_path: str = None,
        conflict_strategy: ConflictStrategy = ConflictStrategy.SKIP,
        progress_callback=None,
        checkpoint: dict = None,
    ) -> ImportResult:
        """从 Excel 导入学生名单，可选照片 ZIP 包

        Args:
            path: Excel 文件路径
            photo_zip_path: 可选照片 ZIP 包路径
            conflict_strategy: 冲突处理策略
            progress_callback: 进度回调函数 callback(current, total)
            checkpoint: 断点续传数据 {'processed_rows': n, 'checkpoint_data': {...}}
        """
        self._cancelled = False
        self._progress_callback = progress_callback

        import zipfile

        from openpyxl import load_workbook

        result = ImportResult()

        # 从断点恢复
        start_row = 0
        if checkpoint and "processed_rows" in checkpoint:
            start_row = checkpoint["processed_rows"]
            if "checkpoint_data" in checkpoint:
                result.checkpoint = checkpoint["checkpoint_data"]

        # 预加载照片映射（如果提供了 ZIP）
        photo_map = {}  # key -> 照片二进制数据
        if photo_zip_path:
            try:
                with zipfile.ZipFile(photo_zip_path, "r") as zf:
                    for name in zf.namelist():
                        if name.lower().endswith((".jpg", ".jpeg", ".png")):
                            # 文件名作为 key（去掉扩展名）
                            key = name.rsplit(".", 1)[0]
                            photo_map[key] = zf.read(name)
            except Exception:
                pass  # 照片加载失败不阻断导入

        result = ImportResult()
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                result.errors.append("文件为空")
                return result
        except Exception as e:
            result.errors.append(f"读取文件失败: {e}")
            return result

        # 自动找标题行(含'姓名'的行)
        headers = [str(v or "").strip() for v in all_rows[0]]
        data_start = 1
        for i, row in enumerate(all_rows):
            if any("姓名" in str(v or "") for v in row):
                headers = [str(v or f"列{j + 1}").strip() for j, v in enumerate(row)]
                data_start = i + 1
                break

        col_map = self._guess_columns(headers)
        if "name" not in col_map:
            result.errors.append("未找到姓名列")
            return result

        # 构建 name→index 映射
        col_index = {}
        for i, h in enumerate(headers):
            h_clean = h.strip().lower().replace(" ", "")
            for key, aliases in self._COLUMN_ALIASES.items():
                for a in aliases:
                    if h_clean == a.strip().lower().replace(" ", ""):
                        col_index[key] = i
                        break

        # 确保用新读取的数据（避免 read_only 模式限制）
        from openpyxl import load_workbook as lw

        wb2 = lw(path, read_only=True)
        ws2 = wb2.active
        all_rows2 = list(ws2.iter_rows(values_only=True))
        wb2.close()

        subjects = {s.name: s.id for s in self.session.query(Subject).all()}
        result.total = len(all_rows2) - data_start

        # 从断点继续
        data_rows = all_rows2[data_start:]
        if start_row >= len(data_rows):
            # 全部已处理完
            return result

        # 更新总数（减去已处理的）
        result.total = len(data_rows) - start_row

        for i, row_data in enumerate(data_rows[start_row:]):
            # 检查取消标志
            if self._cancelled:
                result.checkpoint = {"processed_rows": start_row + i}
                return result

            try:
                name = (
                    str(row_data[col_index["name"]] or "").strip()
                    if col_index.get("name") is not None and col_index["name"] < len(row_data)
                    else ""
                )
                if not name:
                    continue
                cls_raw = (
                    str(row_data[col_index["class_name"]] or "").strip()
                    if "class_name" in col_index and col_index["class_name"] < len(row_data)
                    else ""
                )
                if not cls_raw:
                    result.errors.append(f"{name}: 缺少班级")
                    continue

                cls_name = self._normalize_class_name(cls_raw)
                if not cls_name:
                    result.errors.append(f"{name}: 班级格式无效({cls_raw})")
                    continue

                cls = self.session.query(ClassModel).filter_by(name=cls_name).first()
                if not cls:
                    grade_prefix = cls_name[0]
                    grade_map = {"1": "初一级", "2": "初二级", "3": "初三级"}
                    gn = grade_map.get(grade_prefix)
                    if not gn:
                        result.errors.append(f"{name}: 无法确定年级")
                        continue
                    grade = self.session.query(Grade).filter_by(name=gn).first()
                    if not grade:
                        result.errors.append(f"{name}: 年级不存在({gn})")
                        continue
                    cls = ClassModel(grade_id=grade.id, name=cls_name)
                    self.session.add(cls)
                    self.session.flush()

                # 优先按学籍号匹配（全国学籍号唯一）
                student_code = (
                    str(row_data[col_index.get("student_code", -1)] or "")
                    if "student_code" in col_index and col_index["student_code"] < len(row_data)
                    else ""
                )
                existing = None
                if student_code:
                    existing = (
                        self.session.query(Student).filter_by(student_code=student_code).first()
                    )
                    if existing:
                        # 学籍号匹配到，更新班级（转班情况）
                        if existing.class_id != cls.id:
                            existing.class_id = cls.id
                            result.succeeded += 1
                            continue

                # 次选：按身份证匹配（身份证唯一）
                id_card = (
                    str(row_data[col_index.get("id_card", -1)] or "")
                    if "id_card" in col_index and col_index["id_card"] < len(row_data)
                    else ""
                )
                if id_card:
                    existing = self.session.query(Student).filter_by(id_card=id_card).first()
                    if existing:
                        # 身份证匹配到，更新班级/学籍号（转学/补录情况）
                        if existing.class_id != cls.id:
                            existing.class_id = cls.id
                        if student_code and existing.student_code != student_code:
                            existing.student_code = student_code
                        result.succeeded += 1
                        continue

                # 次选：按班级+姓名匹配
                if not existing:
                    existing = (
                        self.session.query(Student).filter_by(class_id=cls.id, name=name).first()
                    )
                    if existing:
                        # 冲突：同班同名
                        conflict = ConflictInfo(
                            row_index=start_row + i,
                            name=name,
                            student_code=student_code,
                            class_name=cls_name,
                            conflict_type="class_name",
                            existing_student=existing,
                            new_data=self._extract_row_data(row_data, col_index),
                            strategy=ConflictStrategy.SKIP,
                        )
                        result.conflicts.append(conflict)

                        # 根据策略处理
                        if conflict.strategy == ConflictStrategy.OVERWRITE:
                            # 覆盖现有记录
                            old_values = {
                                "class_id": existing.class_id,
                                "name": existing.name,
                                "student_code": existing.student_code,
                            }
                            self._update_student(existing, row_data, col_index, cls.id)
                            # 审计日志
                            manual_audit(
                                self.session,
                                "students",
                                existing.id,
                                "UPDATE",
                                old_values,
                                {"class_id": cls.id},
                            )
                            result.succeeded += 1
                        elif conflict.strategy == ConflictStrategy.KEEP_EXISTING:
                            # 保留现有，跳过
                            result.skipped += 1
                        else:
                            # SKIP - 默认跳过
                            result.skipped += 1
                        continue

                # 再次选：全校同名（无学籍号或学籍号未匹配）
                existing = self.session.query(Student).filter_by(name=name).first()
                if existing:
                    # 冲突：跨班同名
                    conflict = ConflictInfo(
                        row_index=start_row + i,
                        name=name,
                        student_code=student_code,
                        class_name=cls_name,
                        conflict_type="name_cross_class",
                        existing_student=existing,
                        new_data=self._extract_row_data(row_data, col_index),
                        strategy=ConflictStrategy.SKIP,
                    )
                    result.conflicts.append(conflict)

                    if conflict.strategy == ConflictStrategy.OVERWRITE:
                        # 覆盖现有记录（移到新班级）
                        old_values = {
                            "class_id": existing.class_id,
                            "name": existing.name,
                            "student_code": existing.student_code,
                        }
                        existing.class_id = cls.id
                        self._update_student_fields(existing, row_data, col_index)
                        # 审计日志
                        manual_audit(
                            self.session,
                            "students",
                            existing.id,
                            "UPDATE",
                            old_values,
                            {"class_id": cls.id},
                        )
                        result.succeeded += 1
                    elif conflict.strategy == ConflictStrategy.KEEP_EXISTING:
                        result.skipped += 1
                    else:
                        result.skipped += 1
                    continue

                gender = self._clean_gender(
                    str(row_data[col_index.get("gender", -1)] or "")
                    if "gender" in col_index and col_index["gender"] < len(row_data)
                    else ""
                )
                # 确保性别有有效值，默认为"男"
                if not gender:
                    gender = "男"
                student_no = (
                    str(row_data[col_index.get("student_no", -1)] or "")
                    if "student_no" in col_index and col_index["student_no"] < len(row_data)
                    else ""
                )
                id_card = (
                    str(row_data[col_index.get("id_card", -1)] or "")
                    if "id_card" in col_index and col_index["id_card"] < len(row_data)
                    else ""
                )
                phone = (
                    str(row_data[col_index.get("phone", -1)] or "")
                    if "phone" in col_index and col_index["phone"] < len(row_data)
                    else ""
                )
                enroll_year = (
                    self._parse_int(row_data[col_index.get("enroll_year", -1)])
                    if "enroll_year" in col_index and col_index["enroll_year"] < len(row_data)
                    else 0
                )

                # 查找匹配的照片
                photo_data = None
                photo_mime = None
                photo_key_candidates = [
                    student_code,  # 学籍号
                    name,  # 姓名
                    id_card,  # 身份证
                ]
                for key in photo_key_candidates:
                    if key and key in photo_map:
                        photo_data = photo_map[key]
                        photo_mime = (
                            "image/jpeg" if photo_data.startswith(b"\xff\xd8") else "image/png"
                        )
                        break

                student = Student(
                    class_id=cls.id,
                    name=name,
                    student_no=student_no,
                    gender=gender,
                    id_card=id_card,
                    phone=phone,
                    student_code=student_code,
                    enroll_year=enroll_year,
                    photo=photo_data,
                    photo_mime=photo_mime,
                )
                self.session.add(student)
                self.session.flush()  # 确保立即写入数据库，避免后续行重复

                # 审计日志
                manual_audit(
                    self.session,
                    "students",
                    student.id,
                    "CREATE",
                    None,
                    {"name": name, "class_id": cls.id, "student_code": student_code},
                )

                result.succeeded += 1
            except Exception as e:
                result.errors.append(f"{name if 'name' in dir() else '?'}: {e}")

            # 进度回调
            if self._progress_callback:
                self._progress_callback(i + 1, len(data_rows) - start_row)

            # 定期保存断点
            if i % self._checkpoint_interval == 0:
                result.checkpoint = {"processed_rows": start_row + i + 1}

        self.session.commit()
        return result

    def cancel(self):
        """取消导入"""
        self._cancelled = True
        for key, idx in col_index.items():
            if idx < len(row_data):
                data[key] = row_data[idx]
        return data

    def _update_student(self, student: "Student", row_data, col_index: dict, class_id: int):
        """覆盖更新学生记录"""
        student.class_id = class_id
        self._update_student_fields(student, row_data, col_index)

    def _update_student_fields(self, student: "Student", row_data, col_index: dict):
        """更新学生字段"""
        field_map = {
            "name": "name",
            "student_no": "student_no",
            "gender": "gender",
            "id_card": "id_card",
            "phone": "phone",
            "student_code": "student_code",
            "enroll_year": "enroll_year",
        }
        for key, attr in field_map.items():
            if key in col_index and col_index[key] < len(row_data):
                val = row_data[col_index[key]]
                if val is not None:
                    setattr(student, attr, str(val).strip() if isinstance(val, str) else val)

    # 列名别名映射
    _COLUMN_ALIASES = {
        "name": ["姓名", "名字", "学生姓名", "name"],
        "class_name": ["班级", "班别", "class", "班级名称"],
        "student_no": ["座号", "学号", "编号", "序号", "no"],
        "student_code": ["学籍号", "学籍编号", "全国学籍号", "student_code"],
        "gender": ["性别", "男女"],
        "id_card": ["身份证", "证件号", "身份证号"],
        "phone": ["电话", "手机", "联系电话", "phone"],
        "enroll_year": ["入学年份", "入学年", "enroll"],
    }

    def _guess_columns(self, headers: list) -> dict:
        """根据表头猜测列名映射"""
        mapping = {}
        for std_name, aliases in self._COLUMN_ALIASES.items():
            for h in headers:
                h_clean = h.strip().lower().replace(" ", "")
                for a in aliases:
                    if h_clean == a.lower().replace(" ", ""):
                        mapping[std_name] = h
                        break
        return mapping

    def _normalize_class_name(self, raw: str) -> str:
        """班级名称标准化: 初三(1)班 → 301, 3年1班 → 301, 初中2025级2班 → 102"""
        import re

        s = raw.strip()
        # 直接3位数字
        if re.match(r"^\d{3}$", s):
            return s
        # 初三(1)班 / 初三1班
        m = re.search(r"初?([一二三])[\s]*(?:年级?)?[\s(（]*(\d+)[\s)）]*班?", s)
        if m:
            cn_num = {"一": "1", "二": "2", "三": "3"}
            g = cn_num.get(m.group(1), m.group(1))
            return g + m.group(2).zfill(2)[:2]
        # 3年1班
        m = re.search(r"(\d)\s*年\s*(\d+)\s*班?", s)
        if m:
            return m.group(1) + m.group(2).zfill(2)[:2]
        # 初中2025级2班 / 初中2024级1班 / 初中2023级6班
        m = re.search(r"初中(\d{4})级(\d+)班", s)
        if m:
            year = int(m.group(1))
            # 2025 -> 初一(1), 2024 -> 初二(2), 2023 -> 初三(3)
            grade_map = {2025: "1", 2024: "2", 2023: "3"}
            g = grade_map.get(year)
            if g:
                return g + m.group(2).zfill(2)[:2]
        # 纯数字提取
        nums = re.findall(r"\d+", s)
        if nums:
            return "".join(nums)[:3].zfill(3)
        return ""

    def _clean_gender(self, v: str) -> str:
        v_low = v.lower()
        if v_low in ("男", "1", "m", "male"):
            return "男"
        if v_low in ("女", "0", "f", "female"):
            return "女"
        return ""

    def _parse_int(self, v) -> int:
        try:
            return int(float(str(v)))
        except (ValueError, TypeError):
            return 0

    def _extract_row_data(self, row_data, col_index: dict) -> dict:
        data = {}
        for key, idx in col_index.items():
            if idx < len(row_data):
                data[key] = str(row_data[idx] or "").strip()
        return data

    def _update_student(self, student: "Student", row_data, col_index: dict, class_id: int):
        """覆盖更新学生记录"""
        student.class_id = class_id
        self._update_student_fields(student, row_data, col_index)

    def _update_student_fields(self, student: "Student", row_data, col_index: dict):
        """更新学生字段"""
        field_map = {
            "name": "name",
            "student_no": "student_no",
            "gender": "gender",
            "id_card": "id_card",
            "phone": "phone",
            "student_code": "student_code",
            "enroll_year": "enroll_year",
        }
        for key, attr in field_map.items():
            if key in col_index and col_index[key] < len(row_data):
                val = row_data[col_index[key]]
                if val is not None:
                    setattr(student, attr, str(val).strip() if isinstance(val, str) else val)

    # 列名别名映射
    _COLUMN_ALIASES = {
        "name": ["姓名", "名字", "学生姓名", "name"],
        "class_name": ["班级", "班别", "class", "班级名称"],
        "student_no": ["座号", "学号", "编号", "序号", "no"],
        "student_code": ["学籍号", "学籍编号", "全国学籍号", "student_code"],
        "gender": ["性别", "男女"],
        "id_card": ["身份证", "证件号", "身份证号"],
        "phone": ["电话", "手机", "联系电话", "phone"],
        "enroll_year": ["入学年份", "入学年", "enroll"],
    }

    def _guess_columns(self, headers: list) -> dict:
        """根据表头猜测列名映射"""
        mapping = {}
        for std_name, aliases in self._COLUMN_ALIASES.items():
            for h in headers:
                h_clean = h.strip().lower().replace(" ", "")
                for a in aliases:
                    if h_clean == a.lower().replace(" ", ""):
                        mapping[std_name] = h
                        break
        return mapping

    def _normalize_class_name(self, raw: str) -> str:
        """班级名称标准化: 初三(1)班 → 301, 3年1班 → 301, 初中2025级2班 → 102"""
        import re

        s = raw.strip()
        # 直接3位数字
        if re.match(r"^\d{3}$", s):
            return s
        # 初三(1)班 / 初三1班
        m = re.search(r"初?([一二三])[\s]*(?:年级?)?[\s(（]*(\d+)[\s)）]*班?", s)
        if m:
            cn_num = {"一": "1", "二": "2", "三": "3"}
            g = cn_num.get(m.group(1), m.group(1))
            return g + m.group(2).zfill(2)[:2]
        # 3年1班
        m = re.search(r"(\d)\s*年\s*(\d+)\s*班?", s)
        if m:
            return m.group(1) + m.group(2).zfill(2)[:2]
        # 初中2025级2班 / 初中2024级1班 / 初中2023级6班
        m = re.search(r"初中(\d{4})级(\d+)班", s)
        if m:
            year = int(m.group(1))
            # 2025 -> 初一(1), 2024 -> 初二(2), 2023 -> 初三(3)
            grade_map = {2025: "1", 2024: "2", 2023: "3"}
            g = grade_map.get(year)
            if g:
                return g + m.group(2).zfill(2)[:2]
        # 纯数字提取
        nums = re.findall(r"\d+", s)
        if nums:
            return "".join(nums)[:3].zfill(3)
        return ""

    def _clean_gender(self, v: str) -> str:
        v_low = v.lower()
        if v_low in ("男", "1", "m", "male"):
            return "男"
        if v_low in ("女", "0", "f", "female"):
            return "女"
        return ""

    def _parse_int(self, v) -> int:
        try:
            return int(float(str(v)))
        except (ValueError, TypeError):
            return 0

    def _extract_row_data(self, row_data, col_index: dict) -> dict:
        data = {}
        for key, idx in col_index.items():
            if idx < len(row_data):
                data[key] = str(row_data[idx] or "").strip()
        return data

    def _update_student(self, student: "Student", row_data, col_index: dict, class_id: int):
        """覆盖更新学生记录"""
        student.class_id = class_id
        self._update_student_fields(student, row_data, col_index)

    def _update_student_fields(self, student: "Student", row_data, col_index: dict):
        """更新学生字段"""
        field_map = {
            "name": "name",
            "student_no": "student_no",
            "gender": "gender",
            "id_card": "id_card",
            "phone": "phone",
            "student_code": "student_code",
            "enroll_year": "enroll_year",
        }
        for key, attr in field_map.items():
            if key in col_index and col_index[key] < len(row_data):
                val = row_data[col_index[key]]
                if val is not None:
                    setattr(student, attr, str(val).strip() if isinstance(val, str) else val)

    # 列名别名映射
    _COLUMN_ALIASES = {
        "name": ["姓名", "名字", "学生姓名", "name"],
        "class_name": ["班级", "班别", "class", "班级名称"],
        "student_no": ["座号", "学号", "编号", "序号", "no"],
        "student_code": ["学籍号", "学籍编号", "全国学籍号", "student_code"],
        "gender": ["性别", "男女"],
        "id_card": ["身份证", "证件号", "身份证号"],
        "phone": ["电话", "手机", "联系电话", "phone"],
        "enroll_year": ["入学年份", "入学年", "enroll"],
    }

    def _guess_columns(self, headers: list) -> dict:
        """根据表头猜测列名映射"""
        mapping = {}
        for std_name, aliases in self._COLUMN_ALIASES.items():
            for h in headers:
                h_clean = h.strip().lower().replace(" ", "")
                for a in aliases:
                    if h_clean == a.lower().replace(" ", ""):
                        mapping[std_name] = h
                        break
        return mapping

    def _normalize_class_name(self, raw: str) -> str:
        """班级名称标准化: 初三(1)班 → 301, 3年1班 → 301, 初中2025级2班 → 102"""
        import re

        s = raw.strip()
        # 直接3位数字
        if re.match(r"^\d{3}$", s):
            return s
        # 初三(1)班 / 初三1班
        m = re.search(r"初?([一二三])[\s]*(?:年级?)?[\s(（]*(\d+)[\s)）]*班?", s)
        if m:
            cn_num = {"一": "1", "二": "2", "三": "3"}
            g = cn_num.get(m.group(1), m.group(1))
            return g + m.group(2).zfill(2)[:2]
        # 3年1班
        m = re.search(r"(\d)\s*年\s*(\d+)\s*班?", s)
        if m:
            return m.group(1) + m.group(2).zfill(2)[:2]
        # 初中2025级2班 / 初中2024级1班 / 初中2023级6班
        m = re.search(r"初中(\d{4})级(\d+)班", s)
        if m:
            year = int(m.group(1))
            # 2025 -> 初一(1), 2024 -> 初二(2), 2023 -> 初三(3)
            grade_map = {2025: "1", 2024: "2", 2023: "3"}
            g = grade_map.get(year)
            if g:
                return g + m.group(2).zfill(2)[:2]
        # 纯数字提取
        nums = re.findall(r"\d+", s)
        if nums:
            return "".join(nums)[:3].zfill(3)
        return ""

    def _clean_gender(self, v: str) -> str:
        v_low = v.lower()
        if v_low in ("男", "1", "m", "male"):
            return "男"
        if v_low in ("女", "0", "f", "female"):
            return "女"
        return ""

    def _parse_int(self, v) -> int:
        try:
            return int(float(str(v)))
        except (ValueError, TypeError):
            return 0

    def _extract_row_data(self, row_data, col_index: dict) -> dict:
        data = {}
        for key, idx in col_index.items():
            if idx < len(row_data):
                data[key] = str(row_data[idx] or "").strip()
        return data

    def _update_student(self, student: "Student", row_data, col_index: dict, class_id: int):
        """覆盖更新学生记录"""
        student.class_id = class_id
        self._update_student_fields(student, row_data, col_index)

    def _update_student_fields(self, student: "Student", row_data, col_index: dict):
        """更新学生字段"""
        field_map = {
            "name": "name",
            "student_no": "student_no",
            "gender": "gender",
            "id_card": "id_card",
            "phone": "phone",
            "student_code": "student_code",
            "enroll_year": "enroll_year",
        }
        for key, attr in field_map.items():
            if key in col_index and col_index[key] < len(row_data):
                val = row_data[col_index[key]]
                if val is not None:
                    setattr(student, attr, str(val).strip() if isinstance(val, str) else val)

    # 列名别名映射
    _COLUMN_ALIASES = {
        "name": ["姓名", "名字", "学生姓名", "name"],
        "class_name": ["班级", "班别", "class", "班级名称"],
        "student_no": ["座号", "学号", "编号", "序号", "no"],
        "student_code": ["学籍号", "学籍编号", "全国学籍号", "student_code"],
        "gender": ["性别", "男女"],
        "id_card": ["身份证", "证件号", "身份证号"],
        "phone": ["电话", "手机", "联系电话", "phone"],
        "enroll_year": ["入学年份", "入学年", "enroll"],
    }
