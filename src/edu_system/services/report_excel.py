"""
Excel 模板报表服务（Sprint 4.8.1）

- openpyxl 加载 .xlsx 模板 → 按定义名称/占位符填充 → 保留样式与公式
- 模板零侵入：只改单元格值，不动样式/公式/合并单元格
- 支持：单值填充（{{key}} 占位符）、行列表格展开（{{#rows}}...{{/rows}} 区块）
"""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


class ExcelTemplateError(Exception):
    """模板填充错误"""


class ExcelTemplateService:
    def __init__(self):
        pass

    # ── 单值填充 ──
    @staticmethod
    def fill_cells(ws, data: dict):
        """遍历所有单元格，把 {{key}} 占位符替换为 data[key]"""
        replaced = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    new_value = cell.value
                    for key, val in data.items():
                        placeholder = f"{{{{{key}}}}}"
                        if placeholder in new_value:
                            new_value = new_value.replace(placeholder, str(val))
                    if new_value != cell.value:
                        cell.value = new_value
                        replaced += 1
        return replaced

    # ── 行区块展开 ──
    @staticmethod
    def expand_rows(ws, data_rows: list[dict], start_row: int, template_row: int | None = None):
        """在 start_row 起逐行写入数据行；template_row 指定样式模板行（默认 start_row）

        - data_rows: [{col: value, ...}]，cell 用 data_rows[i][列号]
        - 列号映射：cell.column_letter → 同一模板行该列的值会被替换为行数据
        """
        tpl_row = template_row or start_row
        # 读取模板行的占位符结构（cell 坐标 → 占位符 key）
        placeholders: dict[tuple[int, int], str] = {}
        for cell in ws[tpl_row]:
            if isinstance(cell.value, str) and "{{" in cell.value:
                placeholders[(cell.row, cell.column)] = cell.value

        written = 0
        for offset, row_data in enumerate(data_rows):
            target = start_row + offset
            for (trow, tcol), placeholder in placeholders.items():
                # 从占位符提取 key（{{#rows}} 区块内为 {{field}}）
                key = placeholder.strip("{}").strip()
                if key.startswith(("#", "/")):
                    continue
                value = row_data.get(key, "")
                cell = ws.cell(row=target, column=tcol)
                cell.value = value
                # 复制模板行样式
                tpl_cell = ws.cell(row=tpl_row, column=tcol)
                if tpl_cell.has_style:
                    cell.font = copy(tpl_cell.font)
                    cell.border = copy(tpl_cell.border)
                    cell.fill = copy(tpl_cell.fill)
                    cell.alignment = copy(tpl_cell.alignment)
            written += 1
        return written

    # ── 主入口 ──
    @staticmethod
    def render(
        template_path: str | Path,
        data: dict,
        rows: list[dict] | None = None,
        rows_start: int | None = None,
        output_path: str | Path | None = None,
    ) -> bytes:
        """渲染模板：单值填充 + 可选行区块展开

        - data: {{key}} 单值
        - rows/rows_start: 行列表格（rows_start 起）
        - output_path: 给出则保存；否则返回 bytes
        """
        wb = load_workbook(str(template_path))
        ws = wb.active

        ExcelTemplateService.fill_cells(ws, data)
        if rows is not None and rows_start is not None:
            ExcelTemplateService.expand_rows(ws, rows, rows_start)

        if output_path is not None:
            wb.save(str(output_path))
            return Path(output_path).read_bytes()

        import io

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
