"""
报表模板管理服务（M5-D5）

能力：
- 模板注册：登记名称/类型/文件路径（上传后登记）
- 版本管理：同名新模板版本 +1，旧版本保留（可回滚）
- 变量扫描：解析模板文件中的占位符 → 变量列表（{{key}} 或 {{name}} 风格）
- 测试渲染：用样例数据渲染模板，验证变量可用
"""

import json
import re
from pathlib import Path

from sqlalchemy.orm import Session

from edu_system.models import ReportTemplate


class ReportTemplateError(Exception):
    """模板管理异常"""


class ReportTemplateService:
    """报表模板管理服务"""

    # 变量占位符匹配（Excel {{key}} / docxtpl {{name}} 通用）
    PLACEHOLDER_RE = re.compile(r"\{\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*\}\}")

    def __init__(self, session: Session):
        self.session = session

    # ===== 模板注册与版本 =====

    def register(
        self,
        name: str,
        template_type: str,
        file_path: str,
        description: str = "",
        created_by: str = "system",
        variables: list | None = None,
    ) -> ReportTemplate:
        """注册模板。同名已存在 → 版本 +1（历史版本保留）。"""
        if not name or not file_path:
            raise ReportTemplateError("模板名称与文件路径必填")
        if template_type not in ("excel", "word", "certificate"):
            raise ReportTemplateError(f"无效模板类型: {template_type}")

        # 同名模板最新版本
        latest = self.get_latest(name)
        version = (latest.version + 1) if latest else 1
        if latest:
            # 旧版本停用（保持历史）
            latest.is_active = False

        tpl = ReportTemplate(
            name=name,
            template_type=template_type,
            file_path=file_path,
            version=version,
            description=description,
            is_active=True,
            created_by=created_by,
            variables=json.dumps(variables or [], ensure_ascii=False),
        )
        self.session.add(tpl)
        self.session.commit()
        return tpl

    def get_latest(self, name: str) -> ReportTemplate | None:
        """获取同名模板最新版本"""
        return (
            self.session.query(ReportTemplate)
            .filter(ReportTemplate.name == name)
            .order_by(ReportTemplate.version.desc())
            .first()
        )

    def get_versions(self, name: str) -> list[dict]:
        """获取模板全部版本（新→旧）"""
        rows = (
            self.session.query(ReportTemplate)
            .filter(ReportTemplate.name == name)
            .order_by(ReportTemplate.version.desc())
            .all()
        )
        return [
            {
                "version": r.version,
                "file_path": r.file_path,
                "is_active": r.is_active,
                "description": r.description,
                "created_by": r.created_by,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rows
        ]

    def rollback_to(self, name: str, version: int) -> ReportTemplate:
        """回滚到指定版本：该版本置为 active，当前 active 版本停用"""
        target = (
            self.session.query(ReportTemplate)
            .filter(ReportTemplate.name == name, ReportTemplate.version == version)
            .first()
        )
        if not target:
            raise ReportTemplateError(f"版本 {version} 不存在")
        # 停用当前 active
        self.session.query(ReportTemplate).filter(
            ReportTemplate.name == name, ReportTemplate.is_active.is_(True)
        ).update({ReportTemplate.is_active: False})
        target.is_active = True
        self.session.commit()
        return target

    def list_all(self) -> list[dict]:
        """列出所有模板（每模板最新版本）"""
        # 取每组 name 的最大 version
        from sqlalchemy import func

        rows = (
            self.session.query(
                ReportTemplate.name,
                func.max(ReportTemplate.version).label("max_v"),
            )
            .group_by(ReportTemplate.name)
            .all()
        )
        result = []
        for name, max_v in rows:
            tpl = (
                self.session.query(ReportTemplate)
                .filter(
                    ReportTemplate.name == name,
                    ReportTemplate.version == max_v,
                )
                .first()
            )
            result.append(self._to_dict(tpl))
        return result

    def _to_dict(self, tpl: ReportTemplate) -> dict:
        return {
            "id": tpl.id,
            "name": tpl.name,
            "template_type": tpl.template_type,
            "file_path": tpl.file_path,
            "version": tpl.version,
            "is_active": tpl.is_active,
            "description": tpl.description,
            "variables": json.loads(tpl.variables or "[]"),
        }

    # ===== 变量扫描与测试渲染 =====

    def scan_variables(self, template_path: str) -> list[dict]:
        """扫描模板文件中的占位符，返回变量列表 [{key, label}]"""
        p = Path(template_path)
        if not p.exists():
            raise ReportTemplateError(f"模板文件不存在: {template_path}")
        suffix = p.suffix.lower()

        if suffix in (".xlsx", ".xlsm"):
            return self._scan_excel(p)
        if suffix in (".docx",):
            return self._scan_word(p)
        raise ReportTemplateError(f"不支持的模板格式: {suffix}")

    def _scan_excel(self, p: Path) -> list[dict]:
        """扫描 Excel 模板的 {{key}} 占位符"""
        from openpyxl import load_workbook

        wb = load_workbook(p, data_only=False)
        keys: dict[str, str] = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for m in self.PLACEHOLDER_RE.finditer(cell.value):
                            keys.setdefault(m.group(1), m.group(1))
        wb.close()
        return [{"key": k, "label": k} for k in sorted(keys)]

    def _scan_word(self, p: Path) -> list[dict]:
        """扫描 Word 模板的 {{name}} 占位符"""
        import zipfile

        keys: dict[str, str] = {}
        with zipfile.ZipFile(p) as z:
            if "word/document.xml" not in z.namelist():
                return []
            xml_text = z.read("word/document.xml").decode("utf-8", errors="ignore")
            for m in self.PLACEHOLDER_RE.finditer(xml_text):
                keys.setdefault(m.group(1), m.group(1))
        return [{"key": k, "label": k} for k in sorted(keys)]

    def test_render(self, template_path: str, sample_data: dict | None = None) -> dict:
        """测试渲染：用样例数据渲染模板，返回渲染后文本/单元格数"""
        p = Path(template_path)
        if not p.exists():
            raise ReportTemplateError(f"模板文件不存在: {template_path}")
        data = sample_data or {}

        if p.suffix.lower() in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            wb = load_workbook(p, data_only=False)
            rendered_cells = 0
            missing_vars = set()
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            new_val = cell.value
                            for m in self.PLACEHOLDER_RE.finditer(new_val):
                                key = m.group(1)
                                if key in data:
                                    new_val = new_val.replace(m.group(0), str(data[key]))
                                    rendered_cells += 1
                                else:
                                    missing_vars.add(key)
            wb.close()
            return {
                "ok": not missing_vars,
                "rendered_cells": rendered_cells,
                "missing_vars": sorted(missing_vars),
            }

        if p.suffix.lower() in (".docx",):
            import zipfile

            with zipfile.ZipFile(p) as z:
                xml_text = z.read("word/document.xml").decode("utf-8", errors="ignore")
            rendered = 0
            missing = set()
            for m in self.PLACEHOLDER_RE.finditer(xml_text):
                key = m.group(1)
                if key in data:
                    rendered += 1
                else:
                    missing.add(key)
            return {
                "ok": not missing,
                "rendered_cells": rendered,
                "missing_vars": sorted(missing),
            }

        raise ReportTemplateError(f"不支持的模板格式: {p.suffix}")


def get_report_template_service(session: Session = None) -> ReportTemplateService:
    """获取模板服务实例"""
    if session is None:
        from edu_system.database import get_session

        session = get_session()
    return ReportTemplateService(session)
